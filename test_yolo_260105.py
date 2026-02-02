# -*- coding: utf-8 -*-
"""
PDF 컬러 텍스트 추출기 - GUI 버전
PyQt6 기반 데스크톱 애플리케이션
"""
from __future__ import annotations
from pathlib import Path
from dataclasses import dataclass
from typing import List, Tuple, Dict, Optional
import sys
import time
import re
import threading

import pandas as pd
import pymupdf as fitz
from PIL import Image
import io

# ============== PyQt6 임포트 ==============
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QFileDialog, QTextEdit, QProgressBar,
    QTableWidget, QTableWidgetItem, QTabWidget, QGroupBox,
    QSpinBox, QDoubleSpinBox, QCheckBox, QLineEdit, QMessageBox,
    QSplitter, QFrame, QHeaderView, QStyle
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QSize
from PyQt6.QtGui import QFont, QColor, QIcon, QPalette

# ============== YOLO 모델 임포트 ==============
try:
    from ultralytics import YOLO
    YOLO_AVAILABLE = True
except ImportError:
    YOLO_AVAILABLE = False

# ============== Excel 스와치 ==============
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


# ============== 유틸리티 함수들 ==============
def sanitize_for_excel(text: str) -> str:
    """Excel에서 허용되지 않는 불법 문자 제거"""
    if not isinstance(text, str):
        return text
    return ILLEGAL_CHARACTERS_RE.sub("", text)
def hex_to_argb(hex_code: str) -> str:
    if not hex_code: return "FFFFFFFF"
    s = hex_code.strip()
    if s.startswith("#"): s = s[1:]
    if len(s) != 6: return "FFFFFFFF"
    return "FF" + s.upper()


def paint_color_swatches(xlsx_path: Path, swatch_col_name: str, header_row: int = 1):
    wb = load_workbook(xlsx_path)
    ws = wb.active
    col_idx = None
    for c in range(1, ws.max_column + 1):
        if (ws.cell(row=header_row, column=c).value or "").strip() == swatch_col_name:
            col_idx = c
            break
    if col_idx is None:
        wb.close()
        return
    for r in range(header_row + 1, ws.max_row + 1):
        hex_val = ws.cell(row=r, column=col_idx).value
        if isinstance(hex_val, str) and hex_val.strip():
            argb = hex_to_argb(hex_val)
            ws.cell(row=r, column=col_idx).fill = PatternFill(
                fill_type="solid", start_color=argb, end_color=argb
            )
    wb.save(xlsx_path)
    wb.close()


def srgb_int_to_rgb8(srgb_int: int) -> Tuple[int, int, int]:
    r, g, b = fitz.sRGB_to_rgb(srgb_int)
    return int(r), int(g), int(b)


def rgb8_to_hex(rgb: Tuple[int, int, int]) -> str:
    r, g, b = rgb
    return f"#{r:02X}{g:02X}{b:02X}"


def is_grayish(rgb: Tuple[int, int, int], tol: int = 12) -> bool:
    r, g, b = rgb
    return abs(r - g) <= tol and abs(g - b) <= tol


def is_blackish(rgb: Tuple[int, int, int], thr: int = 30) -> bool:
    return max(rgb) < thr


def bbox_center(b: Tuple[float, float, float, float]) -> Tuple[float, float]:
    x0, y0, x1, y1 = b
    return ((x0 + x1) / 2.0, (y0 + y1) / 2.0)


def bbox_union(a: Tuple[float, float, float, float],
               b: Tuple[float, float, float, float]) -> Tuple[float, float, float, float]:
    ax0, ay0, ax1, ay1 = a
    bx0, by0, bx1, by1 = b
    return (min(ax0, bx0), min(ay0, by0), max(ax1, bx1), max(ay1, by1))


# ============== 태그 분류 ==============
# 정확한 라인넘버 패턴 (확정)
# 유형1 (영문시작): WW-0001-50-ASB1B02SN62-HE, CWR-0002-50-ACB3B02SN51-HE
# 유형2 (숫자시작): 308-FH-0095-1000-ACB2B02SN56-HE, 312-H-0004-20-ACB2B09SN51-NN
LINE_NUMBER_PATTERN = re.compile(
    r'^([A-Z]{2,3}-\d{4}-\d{2,3}|\d{3}-[A-Z0-9]{1,4}-\d{4}-\d{2,4})-[A-Z0-9]+-[A-Z]{2}$'
)

TAG_PATTERNS = {
    "subsystem_name": re.compile(r"[0-9]{2,4}-[A-Z0-9]{2,}-[0-9A-Z\-]{4,}\s+[A-Za-z].*"),
    "line_no": LINE_NUMBER_PATTERN,  # 정확한 패턴 사용
    "valve": re.compile(r"\b(?:[A-Z]{1,3})-?\d{3,5}[A-Z]?\b"),
    "instr": re.compile(r"\b[A-Z]{1,3}-?\d{1,5}[A-Z]?\b"),
    "special": re.compile(r"\b(?:SPV|EXJ)\s?-?\s?[0-9A-Z\-]{2,}\b"),
}


def classify_tag(text: str) -> str:
    raw = (text or "").strip()
    upper = raw.upper()
    # interface는 "-IF-"가 포함된 경우만
    if "-IF-" in upper:
        return "interface"
    # subsystem_name은 line_no보다 먼저 체크 (line_no 형식 + 공백 + 텍스트)
    if TAG_PATTERNS["subsystem_name"].search(raw):
        return "subsystem_name"
    # line_no는 정확히 전체 매치되어야 함
    if LINE_NUMBER_PATTERN.match(raw):
        return "line_no"
    if TAG_PATTERNS["special"].search(raw):
        return "special"
    # valve는 하이픈이 정확히 1개여야 함
    if TAG_PATTERNS["valve"].search(raw):
        hyphen_count = raw.count("-")
        if hyphen_count == 1:
            return "valve"
        else:
            return "equipment"
    if TAG_PATTERNS["instr"].search(raw):
        return "instrument"

    # 나머지는 equipment 또는 text 구분
    # 하이픈이 없거나, 온점만 있거나, 괄호만 있는 것은 text이므로 None 반환
    if "-" not in raw:
        return None  # 하이픈 없으면 삭제 대상
    if "." in raw and "-" not in raw:
        return None  # 온점만 있고 하이픈 없으면 삭제 대상
    # 괄호만 있고 하이픈 없는 경우도 삭제 대상 (위에서 이미 처리됨)

    # 하이픈이 포함된 경우만 equipment로 분류
    return "equipment"


# ============== 데이터 구조 ==============
@dataclass
class SpanRec:
    page: int
    text: str
    bbox: Tuple[float, float, float, float]
    rgb: Tuple[int, int, int]
    color_hex: str
    type: str
    pdf_name: str = ""


@dataclass
class ComposedTag:
    page: int
    code: str
    number: str
    composed: str
    code_bbox: Tuple[float, float, float, float]
    number_bbox: Optional[Tuple[float, float, float, float]]
    union_bbox: Tuple[float, float, float, float]
    code_hex: str
    num_hex: Optional[str]
    dy: Optional[float]


# ============== 매칭 파라미터 ==============
# 라인넘버 prefix 패턴 (하이픈으로 끝나는 부분 - 병합용)
# 영문시작: WW-0001-50-, CWR-0002-50-
# 숫자시작: 308-FH-0095-1000-, 313-FH-0032-20-
LINE_PREFIX_RE = re.compile(r"^([A-Z]{2,3}|\d{2,4})-[A-Z0-9\-]+-$", re.IGNORECASE)
LINE_MERGE_DY_MAX = 30.0
LINE_MERGE_X_CENTER_TOL = 20.0

DX_TOL_CENTER = 11.0
DY_TOL_CENTER = 18.0
TARGET_DX = -39.1
TARGET_DY = 1.3
DX_TOL = 10.0
DY_TOL = 22.0
EXPANSIONS = [1.0]

CODE_ONLY_RE = re.compile(r"^[A-Z]{1,4}$")
NUMBER_ONLY_RE = re.compile(r"^\d{2,5}[A-Z]{0,2}$")
LETTER_ONLY_RE = re.compile(r"^[A-Z]{1,3}$")

SUFFIX_LINE_TOL = 3.0
SUFFIX_GAP_MAX = 4.0
EXCLUDE_CODES = {"O", "L", "LL"}


# ============== 텍스트 수집 ==============
def collect_colored_spans(page: fitz.Page, exclude_gray: bool = True, pdf_name: str = "") -> List[SpanRec]:
    spans: List[SpanRec] = []
    d = page.get_text("dict")
    for blk in d.get("blocks", []):
        for line in blk.get("lines", []):
            for s in line.get("spans", []):
                text = (s.get("text") or "").strip()
                if not text:
                    continue
                srgb_int = s.get("color")
                if srgb_int is None:
                    continue
                rgb = srgb_int_to_rgb8(srgb_int)
                if exclude_gray and (is_grayish(rgb) or is_blackish(rgb)):
                    continue
                x0, y0, x1, y1 = s["bbox"]
                spans.append(SpanRec(
                    page=page.number + 1, text=text,
                    bbox=(float(x0), float(y0), float(x1), float(y1)),
                    rgb=rgb, color_hex=rgb8_to_hex(rgb),
                    type=classify_tag(text),
                    pdf_name=pdf_name
                ))
    return spans


def merge_multiline_line_numbers(spans: List[SpanRec]) -> List[SpanRec]:
    """
    두 줄로 나뉜 라인넘버 병합 (reference.py 방식 참고)
    - 하이픈으로 끝나는 텍스트를 찾아 다음 행과 병합 시도
    - 병합 후 정확한 LINE_NUMBER_PATTERN에 매치되면 병합 확정 (line_no로 분류)
    """
    if not spans:
        return spans
    groups: Dict[Tuple[int, str], List[SpanRec]] = {}
    for s in spans:
        groups.setdefault((s.page, s.color_hex), []).append(s)

    merged_all: List[SpanRec] = []
    for (_, _), S in groups.items():
        S_sorted = sorted(S, key=lambda s: (s.bbox[1], s.bbox[0]))
        used = set()
        n = len(S_sorted)
        for i, s in enumerate(S_sorted):
            if i in used:
                continue
            text_top = (s.text or "").strip()

            # 하이픈으로 끝나는 경우 병합 시도
            if text_top.endswith("-"):
                scx, scy = bbox_center(s.bbox)
                merged_flag = False
                for j in range(i + 1, n):
                    if j in used:
                        continue
                    t = S_sorted[j]
                    dy = t.bbox[1] - s.bbox[3]
                    if dy < 0:
                        continue
                    if dy > LINE_MERGE_DY_MAX:
                        break
                    tcx, tcy = bbox_center(t.bbox)
                    if abs(tcx - scx) > LINE_MERGE_X_CENTER_TOL:
                        continue
                    text_bottom = (t.text or "").strip()
                    combined = text_top + text_bottom

                    # 병합 후 정확한 LINE_NUMBER_PATTERN에 매치되면 병합 확정
                    if LINE_NUMBER_PATTERN.match(combined):
                        new_bbox = bbox_union(s.bbox, t.bbox)
                        merged_span = SpanRec(
                            page=s.page,
                            text=combined,
                            bbox=new_bbox,
                            rgb=s.rgb,
                            color_hex=s.color_hex,
                            type="line_no",  # 정확히 매치되었으므로 line_no
                            pdf_name=s.pdf_name
                        )
                        merged_all.append(merged_span)
                        used.add(i)
                        used.add(j)
                        merged_flag = True
                        break
                if not merged_flag:
                    merged_all.append(s)
            else:
                merged_all.append(s)
    return merged_all


def _stitch_suffix(number_span: SpanRec,
                   spans_on_page: List[SpanRec]) -> Tuple[str, Tuple[float, float, float, float]]:
    base_text = number_span.text
    x0, y0, x1, y1 = number_span.bbox
    ncx, ncy = bbox_center(number_span.bbox)

    suffixes: List[SpanRec] = []
    for sp in spans_on_page:
        if sp is number_span:
            continue
        if not LETTER_ONLY_RE.match(sp.text):
            continue
        scx, scy = bbox_center(sp.bbox)
        if abs(scy - ncy) > SUFFIX_LINE_TOL:
            continue
        if sp.bbox[0] >= x1 and (sp.bbox[0] - x1) <= SUFFIX_GAP_MAX:
            suffixes.append(sp)

    suffixes.sort(key=lambda s: s.bbox[0])
    stitched = base_text
    stitched_bbox = (x0, y0, x1, y1)
    appended = 0
    for sp in suffixes:
        if appended >= 2:
            break
        stitched += sp.text
        sx0, sy0, sx1, sy1 = sp.bbox
        stitched_bbox = (min(stitched_bbox[0], sx0),
                         min(stitched_bbox[1], sy0),
                         max(stitched_bbox[2], sx1),
                         max(stitched_bbox[3], sy1))
        appended += 1
    return stitched, stitched_bbox


def _pick_by_window(numbers: List[SpanRec],
                    rx0: float, ry0: float, rx1: float, ry1: float,
                    tx: float, ty: float, cy1: float) -> Optional[Tuple[float, float, SpanRec]]:
    cand = []
    for n in numbers:
        ncx, ncy = bbox_center(n.bbox)
        if rx0 <= ncx <= rx1 and ry0 <= ncy <= ry1:
            ny0 = n.bbox[1]
            dist2 = (ncx - tx) ** 2 + (ncy - ty) ** 2
            dy = ny0 - cy1
            cand.append((dist2, dy, n))
    if not cand:
        return None
    cand.sort(key=lambda x: x[0])
    return cand[0]


def compose_vertical_pairs_simple(spans: List[SpanRec]) -> List[ComposedTag]:
    comps: List[ComposedTag] = []
    spans_by_page: Dict[int, List[SpanRec]] = {}
    for sp in spans:
        spans_by_page.setdefault(sp.page, []).append(sp)

    for page, S in spans_by_page.items():
        codes = [s for s in S
                 if CODE_ONLY_RE.match(s.text)
                 and s.text not in EXCLUDE_CODES]
        numbers = [s for s in S if NUMBER_ONLY_RE.match(s.text)]

        for c in codes:
            cx0, cy0, cx1, cy1 = c.bbox
            ccx, _ = bbox_center(c.bbox)

            chosen = None

            tx = ccx
            ty = cy1 + 0.0
            win = (tx - DX_TOL_CENTER, ty - DY_TOL_CENTER,
                   tx + DX_TOL_CENTER, ty + DY_TOL_CENTER)
            chosen = _pick_by_window(numbers, *win, tx=tx, ty=ty, cy1=cy1)

            if not chosen:
                tx = ccx + TARGET_DX
                ty = cy1 + TARGET_DY
                for scale in EXPANSIONS:
                    win = (tx - DX_TOL * scale, ty - DY_TOL * scale,
                           tx + DX_TOL * scale, ty + DY_TOL * scale)
                    chosen = _pick_by_window(numbers, *win, tx=tx, ty=ty, cy1=cy1)
                    if chosen:
                        break

            if chosen:
                _, dy, n = chosen
                stitched_text, stitched_bbox = _stitch_suffix(n, S)
                comps.append(ComposedTag(
                    page=page,
                    code=c.text,
                    number=stitched_text,
                    composed=f"{c.text}-{stitched_text}",
                    code_bbox=c.bbox,
                    number_bbox=stitched_bbox,
                    union_bbox=bbox_union(c.bbox, stitched_bbox),
                    code_hex=c.color_hex,
                    num_hex=n.color_hex,
                    dy=dy
                ))
            else:
                comps.append(ComposedTag(
                    page=page,
                    code=c.text,
                    number="",
                    composed=c.text,
                    code_bbox=c.bbox,
                    number_bbox=None,
                    union_bbox=c.bbox,
                    code_hex=c.color_hex,
                    num_hex=None,
                    dy=None
                ))
    return comps


# ============== DataFrame 변환 ==============
def to_dataframe(spans: List[SpanRec], comps: List[ComposedTag]):
    df_spans = pd.DataFrame([{
        "page": s.page, "tag": s.text, "type": s.type,
        "x1": s.bbox[0], "y1": s.bbox[1], "x2": s.bbox[2], "y2": s.bbox[3],
        "rgb": s.rgb, "hex": s.color_hex, "hex_swatch": s.color_hex,
        "pdf_name": s.pdf_name
    } for s in spans])

    df_comp = pd.DataFrame([{
        "page": c.page,
        "code": c.code, "number": c.number, "composed": c.composed,
        "code_x1": c.code_bbox[0], "code_y1": c.code_bbox[1],
        "code_x2": c.code_bbox[2], "code_y2": c.code_bbox[3],
        "num_x1": c.number_bbox[0] if c.number_bbox else None,
        "num_y1": c.number_bbox[1] if c.number_bbox else None,
        "num_x2": c.number_bbox[2] if c.number_bbox else None,
        "num_y2": c.number_bbox[3] if c.number_bbox else None,
        "u_x1": c.union_bbox[0], "u_y1": c.union_bbox[1],
        "u_x2": c.union_bbox[2], "u_y2": c.union_bbox[3],
        "code_hex": c.code_hex, "num_hex": c.num_hex,
        "code_hex_swatch": c.code_hex, "num_hex_swatch": c.num_hex,
        "dy": c.dy
    } for c in comps])

    # colored_tags에 composed 반영 (bbox 기반 1:1 매칭)
    if not df_spans.empty and not df_comp.empty:
        df_comp_nonempty = df_comp[df_comp["number"].astype(str) != ""]
        for _, row in df_comp_nonempty.iterrows():
            page = row["page"]
            code = row["code"]
            composed = row["composed"]
            code_hex = row["code_hex"]
            code_x1, code_y1 = row["code_x1"], row["code_y1"]
            code_x2, code_y2 = row["code_x2"], row["code_y2"]

            # bbox 위치까지 매칭하여 정확한 1:1 교체 (tolerance 1.0)
            tol = 1.0
            mask = (
                (df_spans["page"] == page) &
                (df_spans["tag"] == code) &
                (df_spans["hex"] == code_hex) &
                ((df_spans["x1"] - code_x1).abs() <= tol) &
                ((df_spans["y1"] - code_y1).abs() <= tol) &
                ((df_spans["x2"] - code_x2).abs() <= tol) &
                ((df_spans["y2"] - code_y2).abs() <= tol)
            )
            if not mask.any():
                continue

            df_spans.loc[mask, "tag"] = composed
            df_spans.loc[mask, "type"] = "instrument"
            df_spans.loc[mask, ["x1", "y1", "x2", "y2"]] = [
                row["u_x1"], row["u_y1"], row["u_x2"], row["u_y2"]
            ]

    # 불필요 태그 제거
    if not df_spans.empty:
        df_spans["tag"] = df_spans["tag"].astype(str)
        # None 타입 제거 (하이픈 없는 text 등)
        mask_none_type = df_spans["type"].isna()
        mask_exclude_codes = df_spans["tag"].isin(EXCLUDE_CODES)
        mask_digits_only = df_spans["tag"].str.fullmatch(r"\d+")
        upper_tags = df_spans["tag"].str.upper()
        mask_alnum_word = upper_tags.str.fullmatch(r'(?=.*[A-Z])(?=.*\d)[A-Z0-9]+')
        # equipment 타입은 유지 (기존 text 대체)

        df_spans = df_spans[~(mask_none_type | mask_exclude_codes | mask_digits_only | mask_alnum_word)].copy()

        # ============== instrument 후처리 정제 ==============
        if not df_spans.empty:
            # 1. instrument 중 잘못된 접두사로 시작하는 행 삭제
            invalid_prefixes = ("HH-", "HHH-", "STOP-", "OPEN-", "M-", "CLOSE-","DOWN-","UP-")
            mask_instr = df_spans["type"] == "instrument"
            mask_invalid_prefix = df_spans["tag"].str.upper().str.startswith(invalid_prefixes)
            df_spans = df_spans[~(mask_instr & mask_invalid_prefix)].copy()

            # 2. instrument 태그 끝의 불필요 접미사 제거 (긴 것부터 순서대로)
            suffix_pattern = re.compile(r"(MCS|LFO|HHH|LXV|XZV|PSV|HZS|STT|HH|LL|PG|XV|TT|H|L|I|T|E|MSC)$")
            mask_instr = df_spans["type"] == "instrument"
            df_spans.loc[mask_instr, "tag"] = df_spans.loc[mask_instr, "tag"].apply(
                lambda x: suffix_pattern.sub("", x)
            )

            # 3. valve 타입 중 하이픈이 여러 개면 equipment로 변경
            mask_valve = df_spans["type"] == "valve"
            mask_multi_hyphen = df_spans["tag"].str.count("-") > 1
            df_spans.loc[mask_valve & mask_multi_hyphen, "type"] = "equipment"

        df_spans = df_spans.sort_values(["pdf_name", "page", "tag", "x1", "y1"])

    return df_spans, df_comp


# ============== YOLO 관련 ==============
def load_yolo_model(model_path: Path):
    if not YOLO_AVAILABLE:
        return None
    if not model_path.exists():
        return None
    try:
        model = YOLO(model_path.as_posix())
        return model
    except Exception:
        return None


def crop_bbox_from_pdf(pdf_path: Path, page_num: int, bbox: Tuple[float, float, float, float],
                       margin: int = 10) -> Optional[Image.Image]:
    try:
        import cv2
        import numpy as np

        doc = fitz.open(pdf_path.as_posix())
        page = doc[page_num - 1]

        x1, y1, x2, y2 = bbox
        x1 = max(0, x1 - margin)
        y1 = max(0, y1 - margin)
        x2 = min(page.rect.width, x2 + margin)
        y2 = min(page.rect.height, y2 + margin)

        clip_rect = fitz.Rect(x1, y1, x2, y2)
        pix = page.get_pixmap(clip=clip_rect, matrix=fitz.Matrix(2.0, 2.0))

        img_data = pix.tobytes("png")
        img_pil = Image.open(io.BytesIO(img_data))
        img_array = np.array(img_pil)

        if len(img_array.shape) == 3 and img_array.shape[2] == 3:
            img_cv = cv2.cvtColor(img_array, cv2.COLOR_RGB2BGR)
        else:
            img_cv = img_array

        tile_size = 800
        h, w = img_cv.shape[:2]

        if h < tile_size or w < tile_size:
            img_cv = cv2.copyMakeBorder(
                img_cv,
                0, max(0, tile_size - h),
                0, max(0, tile_size - w),
                cv2.BORDER_CONSTANT,
                value=[255, 255, 255]
            )
        elif h > tile_size or w > tile_size:
            center_y, center_x = h // 2, w // 2
            y1_crop = max(0, center_y - tile_size // 2)
            x1_crop = max(0, center_x - tile_size // 2)
            y2_crop = min(h, y1_crop + tile_size)
            x2_crop = min(w, x1_crop + tile_size)
            img_cv = img_cv[y1_crop:y2_crop, x1_crop:x2_crop]

            h_new, w_new = img_cv.shape[:2]
            if h_new < tile_size or w_new < tile_size:
                img_cv = cv2.copyMakeBorder(
                    img_cv,
                    0, max(0, tile_size - h_new),
                    0, max(0, tile_size - w_new),
                    cv2.BORDER_CONSTANT,
                    value=[255, 255, 255]
                )

        img_rgb = cv2.cvtColor(img_cv, cv2.COLOR_BGR2RGB)
        img_pil_final = Image.fromarray(img_rgb)

        doc.close()
        return img_pil_final

    except Exception:
        return None


def classify_with_yolo(model, img: Image.Image, confidence: float = 0.25) -> bool:
    try:
        results = model(img, conf=confidence, imgsz=800, verbose=False)
        if len(results) == 0 or len(results[0].boxes) == 0:
            return False
        for box in results[0].boxes:
            class_id = int(box.cls[0])
            if class_id == 0:
                return True
        return False
    except Exception:
        return False


def reclassify_instruments_with_yolo(df_spans: pd.DataFrame, pdf_dir: Path,
                                     model, confidence: float = 0.25,
                                     progress_callback=None) -> pd.DataFrame:
    if model is None:
        return df_spans
    if df_spans.empty:
        return df_spans

    instruments = df_spans[df_spans["type"] == "instrument"].copy()
    if instruments.empty:
        return df_spans

    reclassified_indices = []
    total = len(instruments)

    for i, (idx, row) in enumerate(instruments.iterrows()):
        pdf_name = row["pdf_name"]
        pdf_path = pdf_dir / pdf_name

        if not pdf_path.exists():
            continue

        page_num = int(row["page"])
        bbox = (row["x1"], row["y1"], row["x2"], row["y2"])

        img = crop_bbox_from_pdf(pdf_path, page_num, bbox, margin=10)
        if img is None:
            continue

        is_special = classify_with_yolo(model, img, confidence=confidence)
        if is_special:
            reclassified_indices.append(idx)

        if progress_callback:
            progress_callback(i + 1, total)

    if reclassified_indices:
        df_spans.loc[reclassified_indices, "type"] = "special_item"

    return df_spans


# ============== 처리 워커 스레드 ==============
class ProcessWorker(QThread):
    """백그라운드 처리 스레드"""
    progress = pyqtSignal(int, int, str)  # current, total, message
    log = pyqtSignal(str)
    finished_signal = pyqtSignal(object, object)  # df_spans, df_comp
    error = pyqtSignal(str)

    def __init__(self, pdf_files: List[Path], output_dir: Path,
                 use_yolo: bool = False, yolo_path: str = "",
                 yolo_confidence: float = 0.25):
        super().__init__()
        self.pdf_files = pdf_files
        self.output_dir = output_dir
        self.use_yolo = use_yolo
        self.yolo_path = yolo_path
        self.yolo_confidence = yolo_confidence
        self._is_cancelled = False

    def cancel(self):
        self._is_cancelled = True

    def run(self):
        try:
            all_spans: List[SpanRec] = []
            all_comp: List[ComposedTag] = []

            total_files = len(self.pdf_files)

            for file_idx, pdf_path in enumerate(self.pdf_files):
                if self._is_cancelled:
                    self.log.emit("❌ 작업이 취소되었습니다.")
                    return

                self.log.emit(f"📄 처리 중: {pdf_path.name}")
                self.progress.emit(file_idx, total_files, f"PDF 처리: {pdf_path.name}")

                try:
                    doc = fitz.open(pdf_path.as_posix())
                    page_count = doc.page_count

                    for i, page in enumerate(doc):
                        if self._is_cancelled:
                            doc.close()
                            return

                        spans = collect_colored_spans(page, exclude_gray=True, pdf_name=pdf_path.name)
                        spans = merge_multiline_line_numbers(spans)
                        comps = compose_vertical_pairs_simple(spans)
                        all_spans.extend(spans)
                        all_comp.extend(comps)

                        self.log.emit(f"  페이지 {i+1}/{page_count}: spans={len(spans)}, composed={len(comps)}")

                    doc.close()
                    self.log.emit(f"✅ 완료: {pdf_path.name}")

                except Exception as e:
                    self.log.emit(f"⚠️ 오류 ({pdf_path.name}): {str(e)}")
                    continue

            self.progress.emit(total_files, total_files, "DataFrame 변환 중...")
            self.log.emit("📊 DataFrame 변환 중...")

            df_spans, df_comp = to_dataframe(all_spans, all_comp)

            # YOLO 재분류
            if self.use_yolo and self.yolo_path:
                self.log.emit("🔍 YOLO 모델 로드 중...")
                model = load_yolo_model(Path(self.yolo_path))

                if model is not None:
                    self.log.emit(f"✅ YOLO 모델 로드 완료")

                    pdf_dir = self.pdf_files[0].parent if self.pdf_files else Path(".")

                    def yolo_progress(current, total):
                        self.progress.emit(current, total, f"YOLO 분류: {current}/{total}")

                    df_spans = reclassify_instruments_with_yolo(
                        df_spans, pdf_dir, model,
                        confidence=self.yolo_confidence,
                        progress_callback=yolo_progress
                    )

                    special_count = (df_spans["type"] == "special_item").sum()
                    self.log.emit(f"✨ YOLO 재분류 완료: {special_count}개 special_item 발견")
                else:
                    self.log.emit("⚠️ YOLO 모델 로드 실패 - 재분류 건너뜀")

            # 결과 저장
            self.log.emit("💾 결과 저장 중...")
            self.output_dir.mkdir(parents=True, exist_ok=True)

            csv_spans = self.output_dir / "colored_tags.csv"
            xlsx_spans = self.output_dir / "colored_tags.xlsx"
            csv_comp = self.output_dir / "composed_tags.csv"
            xlsx_comp = self.output_dir / "composed_tags.xlsx"

            # Excel 저장 전 불법 문자 제거 (ɆȺɌⱾɊɁȺɌ 등 유니코드 특수문자)
            df_spans_clean = df_spans.copy()
            df_comp_clean = df_comp.copy()
            for col in df_spans_clean.select_dtypes(include=['object']).columns:
                df_spans_clean[col] = df_spans_clean[col].apply(sanitize_for_excel)
            for col in df_comp_clean.select_dtypes(include=['object']).columns:
                df_comp_clean[col] = df_comp_clean[col].apply(sanitize_for_excel)

            df_spans_clean.to_csv(csv_spans, index=False, encoding="utf-8-sig")
            df_comp_clean.to_csv(csv_comp, index=False, encoding="utf-8-sig")
            df_spans_clean.to_excel(xlsx_spans, index=False)
            df_comp_clean.to_excel(xlsx_comp, index=False)

            # Excel 색상 스와치
            for path, col in [(xlsx_spans, "hex_swatch"),
                              (xlsx_comp, "code_hex_swatch"),
                              (xlsx_comp, "num_hex_swatch")]:
                try:
                    paint_color_swatches(path, swatch_col_name=col)
                except Exception:
                    pass

            self.log.emit(f"✅ 저장 완료:")
            self.log.emit(f"   - {csv_spans}")
            self.log.emit(f"   - {xlsx_spans}")
            self.log.emit(f"   - {csv_comp}")
            self.log.emit(f"   - {xlsx_comp}")

            self.finished_signal.emit(df_spans, df_comp)

        except Exception as e:
            self.error.emit(str(e))


# ============== 메인 GUI ==============
class PDFColorExtractorGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.pdf_files: List[Path] = []
        self.output_dir: Path = Path("out")
        self.worker: Optional[ProcessWorker] = None
        self.df_spans = None
        self.df_comp = None

        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("PDF 컬러 텍스트 추출기")
        self.setMinimumSize(1000, 700)

        # 중앙 위젯
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # 상단: 파일 선택
        file_group = QGroupBox("📁 파일 설정")
        file_layout = QVBoxLayout(file_group)

        # PDF 파일 선택
        pdf_row = QHBoxLayout()
        pdf_row.addWidget(QLabel("PDF 파일:"))
        self.pdf_label = QLabel("선택된 파일 없음")
        self.pdf_label.setStyleSheet("color: gray;")
        pdf_row.addWidget(self.pdf_label, 1)
        self.btn_select_pdf = QPushButton("파일 선택")
        self.btn_select_pdf.clicked.connect(self.select_pdf_files)
        pdf_row.addWidget(self.btn_select_pdf)
        self.btn_select_folder = QPushButton("폴더 선택")
        self.btn_select_folder.clicked.connect(self.select_pdf_folder)
        pdf_row.addWidget(self.btn_select_folder)
        file_layout.addLayout(pdf_row)

        # 출력 폴더 선택
        out_row = QHBoxLayout()
        out_row.addWidget(QLabel("출력 폴더:"))
        self.out_label = QLabel(str(self.output_dir.resolve()))
        out_row.addWidget(self.out_label, 1)
        self.btn_select_out = QPushButton("변경")
        self.btn_select_out.clicked.connect(self.select_output_dir)
        out_row.addWidget(self.btn_select_out)
        file_layout.addLayout(out_row)

        main_layout.addWidget(file_group)

        # 중간: 설정
        settings_group = QGroupBox("⚙️ 설정")
        settings_layout = QHBoxLayout(settings_group)

        # YOLO 설정
        yolo_box = QVBoxLayout()
        self.chk_use_yolo = QCheckBox("YOLO 모델 사용 (special_item 분류)")
        self.chk_use_yolo.setChecked(False)
        self.chk_use_yolo.toggled.connect(self.toggle_yolo_settings)
        yolo_box.addWidget(self.chk_use_yolo)

        yolo_path_row = QHBoxLayout()
        yolo_path_row.addWidget(QLabel("모델 경로:"))
        self.yolo_path_edit = QLineEdit()
        self.yolo_path_edit.setPlaceholderText("runs/detect/symbol_detector/weights/best.pt")
        self.yolo_path_edit.setEnabled(False)
        yolo_path_row.addWidget(self.yolo_path_edit, 1)
        self.btn_select_yolo = QPushButton("찾기")
        self.btn_select_yolo.setEnabled(False)
        self.btn_select_yolo.clicked.connect(self.select_yolo_model)
        yolo_path_row.addWidget(self.btn_select_yolo)
        yolo_box.addLayout(yolo_path_row)

        yolo_conf_row = QHBoxLayout()
        yolo_conf_row.addWidget(QLabel("Confidence:"))
        self.yolo_conf_spin = QDoubleSpinBox()
        self.yolo_conf_spin.setRange(0.01, 1.0)
        self.yolo_conf_spin.setSingleStep(0.05)
        self.yolo_conf_spin.setValue(0.25)
        self.yolo_conf_spin.setEnabled(False)
        yolo_conf_row.addWidget(self.yolo_conf_spin)
        yolo_conf_row.addStretch()
        yolo_box.addLayout(yolo_conf_row)

        settings_layout.addLayout(yolo_box)

        # YOLO 상태
        if not YOLO_AVAILABLE:
            yolo_status = QLabel("⚠️ ultralytics 미설치")
            yolo_status.setStyleSheet("color: orange;")
            self.chk_use_yolo.setEnabled(False)
        else:
            yolo_status = QLabel("✅ YOLO 사용 가능")
            yolo_status.setStyleSheet("color: green;")
        settings_layout.addWidget(yolo_status)
        settings_layout.addStretch()

        main_layout.addWidget(settings_group)

        # 실행 버튼
        btn_row = QHBoxLayout()
        self.btn_run = QPushButton("▶️ 실행")
        self.btn_run.setMinimumHeight(40)
        self.btn_run.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-size: 14px;
                font-weight: bold;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
        """)
        self.btn_run.clicked.connect(self.run_extraction)
        btn_row.addWidget(self.btn_run)

        self.btn_cancel = QPushButton("⏹️ 취소")
        self.btn_cancel.setMinimumHeight(40)
        self.btn_cancel.setEnabled(False)
        self.btn_cancel.clicked.connect(self.cancel_extraction)
        btn_row.addWidget(self.btn_cancel)

        main_layout.addLayout(btn_row)

        # 진행 상황
        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(True)
        main_layout.addWidget(self.progress_bar)

        self.status_label = QLabel("대기 중...")
        main_layout.addWidget(self.status_label)

        # 하단: 탭 (로그 + 결과)
        splitter = QSplitter(Qt.Orientation.Vertical)

        # 로그
        log_group = QGroupBox("📝 로그")
        log_layout = QVBoxLayout(log_group)
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setFont(QFont("Consolas", 9))
        log_layout.addWidget(self.log_text)
        splitter.addWidget(log_group)

        # 결과 탭
        result_tabs = QTabWidget()

        # colored_tags 탭
        self.table_spans = QTableWidget()
        self.table_spans.setAlternatingRowColors(True)
        result_tabs.addTab(self.table_spans, "Colored Tags")

        # composed_tags 탭
        self.table_comp = QTableWidget()
        self.table_comp.setAlternatingRowColors(True)
        result_tabs.addTab(self.table_comp, "Composed Tags")

        splitter.addWidget(result_tabs)
        splitter.setSizes([200, 300])

        main_layout.addWidget(splitter, 1)

    def toggle_yolo_settings(self, checked):
        self.yolo_path_edit.setEnabled(checked)
        self.btn_select_yolo.setEnabled(checked)
        self.yolo_conf_spin.setEnabled(checked)

    def select_pdf_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self, "PDF 파일 선택", "",
            "PDF Files (*.pdf);;All Files (*)"
        )
        if files:
            self.pdf_files = [Path(f) for f in files]
            self.pdf_label.setText(f"{len(self.pdf_files)}개 파일 선택됨")
            self.pdf_label.setStyleSheet("color: black;")
            self.log_text.append(f"📄 {len(self.pdf_files)}개 파일 선택:")
            for f in self.pdf_files[:5]:
                self.log_text.append(f"   - {f.name}")
            if len(self.pdf_files) > 5:
                self.log_text.append(f"   ... 외 {len(self.pdf_files)-5}개")

    def select_pdf_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "PDF 폴더 선택")
        if folder:
            folder_path = Path(folder)
            self.pdf_files = sorted(folder_path.glob("*.pdf"))
            if self.pdf_files:
                self.pdf_label.setText(f"{len(self.pdf_files)}개 파일 (폴더: {folder_path.name})")
                self.pdf_label.setStyleSheet("color: black;")
                self.log_text.append(f"📁 폴더 선택: {folder_path}")
                self.log_text.append(f"   {len(self.pdf_files)}개 PDF 파일 발견")
            else:
                self.pdf_label.setText("선택된 폴더에 PDF 없음")
                self.pdf_label.setStyleSheet("color: red;")
                QMessageBox.warning(self, "경고", "선택된 폴더에 PDF 파일이 없습니다.")

    def select_output_dir(self):
        folder = QFileDialog.getExistingDirectory(self, "출력 폴더 선택")
        if folder:
            self.output_dir = Path(folder)
            self.out_label.setText(str(self.output_dir.resolve()))

    def select_yolo_model(self):
        file, _ = QFileDialog.getOpenFileName(
            self, "YOLO 모델 선택", "",
            "PyTorch Model (*.pt);;All Files (*)"
        )
        if file:
            self.yolo_path_edit.setText(file)

    def run_extraction(self):
        if not self.pdf_files:
            QMessageBox.warning(self, "경고", "PDF 파일을 선택해주세요.")
            return

        self.btn_run.setEnabled(False)
        self.btn_cancel.setEnabled(True)
        self.progress_bar.setValue(0)
        self.log_text.clear()
        self.log_text.append("🚀 추출 시작...")

        self.worker = ProcessWorker(
            pdf_files=self.pdf_files,
            output_dir=self.output_dir,
            use_yolo=self.chk_use_yolo.isChecked(),
            yolo_path=self.yolo_path_edit.text(),
            yolo_confidence=self.yolo_conf_spin.value()
        )

        self.worker.progress.connect(self.on_progress)
        self.worker.log.connect(self.on_log)
        self.worker.finished_signal.connect(self.on_finished)
        self.worker.error.connect(self.on_error)

        self.worker.start()

    def cancel_extraction(self):
        if self.worker:
            self.worker.cancel()
            self.btn_cancel.setEnabled(False)
            self.status_label.setText("취소 중...")

    def on_progress(self, current, total, message):
        if total > 0:
            percent = int(current / total * 100)
            self.progress_bar.setValue(percent)
        self.status_label.setText(message)

    def on_log(self, message):
        self.log_text.append(message)
        # 자동 스크롤
        scrollbar = self.log_text.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())

    def on_finished(self, df_spans, df_comp):
        self.btn_run.setEnabled(True)
        self.btn_cancel.setEnabled(False)
        self.progress_bar.setValue(100)
        self.status_label.setText("✅ 완료!")

        self.df_spans = df_spans
        self.df_comp = df_comp

        # 테이블 업데이트
        self.update_table(self.table_spans, df_spans)
        self.update_table(self.table_comp, df_comp)

        self.log_text.append("")
        self.log_text.append("=" * 50)
        self.log_text.append(f"📊 결과 요약:")
        self.log_text.append(f"   - Colored Tags: {len(df_spans)}개")
        self.log_text.append(f"   - Composed Tags: {len(df_comp)}개")

        if not df_spans.empty:
            type_counts = df_spans["type"].value_counts()
            self.log_text.append(f"   - 타입별 분포:")
            for t, c in type_counts.items():
                self.log_text.append(f"      {t}: {c}개")

        QMessageBox.information(self, "완료", f"추출 완료!\n\n결과 저장 위치: {self.output_dir}")

    def on_error(self, error_msg):
        self.btn_run.setEnabled(True)
        self.btn_cancel.setEnabled(False)
        self.status_label.setText("❌ 오류 발생")
        self.log_text.append(f"❌ 오류: {error_msg}")
        QMessageBox.critical(self, "오류", f"처리 중 오류 발생:\n{error_msg}")

    def update_table(self, table: QTableWidget, df: pd.DataFrame):
        if df is None or df.empty:
            table.clear()
            table.setRowCount(0)
            table.setColumnCount(0)
            return

        # 최대 1000행만 표시
        display_df = df.head(1000)

        table.setRowCount(len(display_df))
        table.setColumnCount(len(display_df.columns))
        table.setHorizontalHeaderLabels(display_df.columns.tolist())

        for row_idx, (_, row) in enumerate(display_df.iterrows()):
            for col_idx, value in enumerate(row):
                item = QTableWidgetItem(str(value) if pd.notna(value) else "")

                # hex 컬럼이면 배경색 적용
                col_name = display_df.columns[col_idx]
                if "hex" in col_name.lower() and isinstance(value, str) and value.startswith("#"):
                    try:
                        color = QColor(value)
                        item.setBackground(color)
                        # 밝기에 따라 글자색 조정
                        brightness = (color.red() * 299 + color.green() * 587 + color.blue() * 114) / 1000
                        if brightness < 128:
                            item.setForeground(QColor(255, 255, 255))
                    except:
                        pass

                table.setItem(row_idx, col_idx, item)

        table.resizeColumnsToContents()

        if len(df) > 1000:
            self.log_text.append(f"⚠️ 테이블은 처음 1000행만 표시 (전체: {len(df)}행)")


def main():
    app = QApplication(sys.argv)

    # 스타일 설정
    app.setStyle("Fusion")

    window = PDFColorExtractorGUI()
    window.show()

    sys.exit(app.exec())


if __name__ == "__main__":
    main()