# -*- coding: utf-8 -*-
"""
PDF Annotation Equipment & Instrument 추출기 - GUI 버전
PDF 주석(annotation)에서 Equipment Number와 Instrument Tag 추출
"""
from __future__ import annotations
from pathlib import Path
from dataclasses import dataclass
from typing import List, Tuple, Dict, Optional
import sys
import re

import pandas as pd
import pymupdf as fitz
from PIL import Image
import io

# ============== PyQt6 임포트 ==============
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QFileDialog, QTextEdit, QProgressBar,
    QTableWidget, QTableWidgetItem, QTabWidget, QGroupBox,
    QSplitter, QHeaderView, QMessageBox
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QFont, QColor

# ============== Excel 스타일 ==============
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


# ============== 유틸리티 함수들 ==============
def sanitize_for_excel(text: str) -> str:
    """Excel에서 허용되지 않는 불법 문자 제거"""
    if not isinstance(text, str):
        return text
    return ILLEGAL_CHARACTERS_RE.sub("", text)


def hex_to_argb(hex_code: str) -> str:
    if not hex_code: return "FFFFFFFF"
    s = hex_code.strip()
    if s.startswith("#"): s = s[1:]
    if len(s) != 6: return "FFFFFFFF"
    return "FF" + s.upper()


def paint_color_swatches(xlsx_path: Path, swatch_col_name: str, header_row: int = 1):
    wb = load_workbook(xlsx_path)
    ws = wb.active
    col_idx = None
    for c in range(1, ws.max_column + 1):
        if (ws.cell(row=header_row, column=c).value or "").strip() == swatch_col_name:
            col_idx = c
            break
    if col_idx is None:
        wb.close()
        return
    for r in range(header_row + 1, ws.max_row + 1):
        hex_val = ws.cell(row=r, column=col_idx).value
        if isinstance(hex_val, str) and hex_val.strip():
            argb = hex_to_argb(hex_val)
            ws.cell(row=r, column=col_idx).fill = PatternFill(
                fill_type="solid", start_color=argb, end_color=argb
            )
    wb.save(xlsx_path)
    wb.close()


def rgb_to_hex(rgb: Tuple[float, float, float]) -> str:
    """RGB (0-1 range) to hex string"""
    r = int(rgb[0] * 255)
    g = int(rgb[1] * 255)
    b = int(rgb[2] * 255)
    return f"#{r:02X}{g:02X}{b:02X}"


def rgb8_to_hex(rgb: Tuple[int, int, int]) -> str:
    """RGB (0-255 range) to hex string"""
    r, g, b = rgb
    return f"#{r:02X}{g:02X}{b:02X}"


def is_grayish(rgb: Tuple[int, int, int], tol: int = 12) -> bool:
    r, g, b = rgb
    return abs(r - g) <= tol and abs(g - b) <= tol


def is_blackish(rgb: Tuple[int, int, int], thr: int = 30) -> bool:
    return max(rgb) < thr


def bbox_center(b: Tuple[float, float, float, float]) -> Tuple[float, float]:
    x0, y0, x1, y1 = b
    return ((x0 + x1) / 2.0, (y0 + y1) / 2.0)


def bbox_union(a: Tuple[float, float, float, float],
               b: Tuple[float, float, float, float]) -> Tuple[float, float, float, float]:
    ax0, ay0, ax1, ay1 = a
    bx0, by0, bx1, by1 = b
    return (min(ax0, bx0), min(ay0, by0), max(ax1, bx1), max(ay1, by1))


# ============== 태그 분류 패턴 ==============

# Equipment Number 패턴
# 예: AL-8903A, TK-8901, PT-8903A, LSH-8903A, XV-8903C, PSV-8903A
EQUIPMENT_PATTERN = re.compile(
    r'^[A-Z]{2,4}-\d{4,5}[A-Z]{0,2}$'
)

# Instrument 패턴 (수직 배치로 구성된 태그)
CODE_ONLY_RE = re.compile(r"^[A-Z]{1,4}$")
NUMBER_ONLY_RE = re.compile(r"^\d{4,5}[A-Z]{0,2}$")
LETTER_ONLY_RE = re.compile(r"^[A-Z]{1,3}$")

# 제외할 코드
EXCLUDE_CODES = {"O", "L", "LL", "TO", "FC", "I", "S", "V", "D"}


def classify_tag(text: str) -> str:
    """태그 분류"""
    raw = (text or "").strip()
    upper = raw.upper()

    # 빈 문자열 제외
    if not raw:
        return None

    # Equipment Number 체크
    if EQUIPMENT_PATTERN.match(raw):
        return "equipment"

    # Code only (instrument 구성 요소)
    if CODE_ONLY_RE.match(raw) and raw not in EXCLUDE_CODES:
        return "code_part"

    # Number only (instrument 구성 요소)
    if NUMBER_ONLY_RE.match(raw):
        return "number_part"

    # 그 외는 제외
    return None


# ============== 데이터 구조 ==============
@dataclass
class AnnotRec:
    """주석 레코드"""
    page: int
    text: str
    bbox: Tuple[float, float, float, float]
    rgb: Tuple[int, int, int]
    color_hex: str
    type: str
    pdf_name: str = ""


@dataclass
class ComposedTag:
    page: int
    code: str
    number: str
    composed: str
    code_bbox: Tuple[float, float, float, float]
    number_bbox: Optional[Tuple[float, float, float, float]]
    union_bbox: Tuple[float, float, float, float]
    code_hex: str
    num_hex: Optional[str]
    dy: Optional[float]
    pdf_name: str = ""


# ============== 매칭 파라미터 ==============
DX_TOL_CENTER = 11.0
DY_TOL_CENTER = 18.0
TARGET_DX = -39.1
TARGET_DY = 1.3
DX_TOL = 10.0
DY_TOL = 22.0
EXPANSIONS = [1.0]

SUFFIX_LINE_TOL = 3.0
SUFFIX_GAP_MAX = 4.0


# ============== 주석 수집 ==============
def collect_colored_annotations(page: fitz.Page, exclude_gray: bool = True, pdf_name: str = "") -> List[AnnotRec]:
    """페이지에서 컬러 주석(annotation) 수집"""
    annots: List[AnnotRec] = []

    try:
        # 주석 순회
        for annot in page.annots():
            # 주석 텍스트 추출
            text = ""

            # 주석 내용 추출 시도 (여러 방법)
            if annot.info.get("content"):
                text = annot.info["content"].strip()
            elif annot.info.get("subject"):
                text = annot.info["subject"].strip()

            # FreeText 주석의 경우
            if annot.type[0] == fitz.PDF_ANNOT_FREE_TEXT:
                # get_text() 메서드가 있으면 사용
                try:
                    if hasattr(annot, 'get_text'):
                        extracted = annot.get_text()
                        if extracted:
                            text = extracted.strip()
                except:
                    pass

            # 텍스트가 없으면 건너뛰기
            if not text:
                continue

            # 주석 색상 추출
            colors = annot.colors
            if not colors or "stroke" not in colors:
                # 색상이 없으면 검정으로 간주
                rgb = (0, 0, 0)
            else:
                # stroke 색상 사용 (0-1 범위)
                stroke_color = colors["stroke"]
                if len(stroke_color) >= 3:
                    rgb = (
                        int(stroke_color[0] * 255),
                        int(stroke_color[1] * 255),
                        int(stroke_color[2] * 255)
                    )
                else:
                    rgb = (0, 0, 0)

            # 회색/검은색 제외
            if exclude_gray and (is_grayish(rgb) or is_blackish(rgb)):
                continue

            # bbox 추출
            rect = annot.rect
            bbox = (float(rect.x0), float(rect.y0), float(rect.x1), float(rect.y1))

            # 태그 분류
            tag_type = classify_tag(text)

            # 분류된 태그만 수집
            if tag_type:
                annots.append(AnnotRec(
                    page=page.number + 1,
                    text=text,
                    bbox=bbox,
                    rgb=rgb,
                    color_hex=rgb8_to_hex(rgb),
                    type=tag_type,
                    pdf_name=pdf_name
                ))
    except Exception as e:
        # 주석 처리 중 오류 발생 시 무시
        pass

    return annots


def _stitch_suffix(number_annot: AnnotRec, annots_on_page: List[AnnotRec]) -> Tuple[str, Tuple[float, float, float, float]]:
    """번호 뒤에 붙는 알파벳 접미사 연결 (예: 8903 + A = 8903A)"""
    base_text = number_annot.text
    x0, y0, x1, y1 = number_annot.bbox
    ncx, ncy = bbox_center(number_annot.bbox)

    suffixes: List[AnnotRec] = []
    for an in annots_on_page:
        if an is number_annot:
            continue
        if not LETTER_ONLY_RE.match(an.text):
            continue
        scx, scy = bbox_center(an.bbox)
        if abs(scy - ncy) > SUFFIX_LINE_TOL:
            continue
        if an.bbox[0] >= x1 and (an.bbox[0] - x1) <= SUFFIX_GAP_MAX:
            suffixes.append(an)

    suffixes.sort(key=lambda s: s.bbox[0])
    stitched = base_text
    stitched_bbox = (x0, y0, x1, y1)
    appended = 0

    for an in suffixes:
        if appended >= 2:
            break
        stitched += an.text
        sx0, sy0, sx1, sy1 = an.bbox
        stitched_bbox = (min(stitched_bbox[0], sx0),
                         min(stitched_bbox[1], sy0),
                         max(stitched_bbox[2], sx1),
                         max(stitched_bbox[3], sy1))
        appended += 1

    return stitched, stitched_bbox


def _pick_by_window(numbers: List[AnnotRec],
                    rx0: float, ry0: float, rx1: float, ry1: float,
                    tx: float, ty: float, cy1: float) -> Optional[Tuple[float, float, AnnotRec]]:
    """윈도우 내에서 가장 가까운 번호 선택"""
    cand = []
    for n in numbers:
        ncx, ncy = bbox_center(n.bbox)
        if rx0 <= ncx <= rx1 and ry0 <= ncy <= ry1:
            ny0 = n.bbox[1]
            dist2 = (ncx - tx) ** 2 + (ncy - ty) ** 2
            dy = ny0 - cy1
            cand.append((dist2, dy, n))

    if not cand:
        return None

    cand.sort(key=lambda x: x[0])
    return cand[0]


def compose_vertical_pairs(annots: List[AnnotRec]) -> List[ComposedTag]:
    """수직으로 배치된 코드-번호 쌍 구성 (Instrument)"""
    comps: List[ComposedTag] = []
    annots_by_page: Dict[int, List[AnnotRec]] = {}

    for an in annots:
        annots_by_page.setdefault(an.page, []).append(an)

    for page, A in annots_by_page.items():
        # 코드 부분만 추출
        codes = [a for a in A if a.type == "code_part"]
        # 번호 부분만 추출
        numbers = [a for a in A if a.type == "number_part"]

        # 페이지의 pdf_name 가져오기 (첫 번째 annotation의 pdf_name 사용)
        pdf_name = A[0].pdf_name if A else ""

        for c in codes:
            cx0, cy0, cx1, cy1 = c.bbox
            ccx, _ = bbox_center(c.bbox)

            chosen = None

            # 1단계: 중앙 아래 탐색
            tx = ccx
            ty = cy1 + 0.0
            win = (tx - DX_TOL_CENTER, ty - DY_TOL_CENTER,
                   tx + DX_TOL_CENTER, ty + DY_TOL_CENTER)
            chosen = _pick_by_window(numbers, *win, tx=tx, ty=ty, cy1=cy1)

            # 2단계: 타겟 위치 탐색
            if not chosen:
                tx = ccx + TARGET_DX
                ty = cy1 + TARGET_DY
                for scale in EXPANSIONS:
                    win = (tx - DX_TOL * scale, ty - DY_TOL * scale,
                           tx + DX_TOL * scale, ty + DY_TOL * scale)
                    chosen = _pick_by_window(numbers, *win, tx=tx, ty=ty, cy1=cy1)
                    if chosen:
                        break

            if chosen:
                _, dy, n = chosen
                stitched_text, stitched_bbox = _stitch_suffix(n, A)

                comps.append(ComposedTag(
                    page=page,
                    code=c.text,
                    number=stitched_text,
                    composed=f"{c.text}-{stitched_text}",
                    code_bbox=c.bbox,
                    number_bbox=stitched_bbox,
                    union_bbox=bbox_union(c.bbox, stitched_bbox),
                    code_hex=c.color_hex,
                    num_hex=n.color_hex,
                    dy=dy,
                    pdf_name=pdf_name
                ))

    return comps


# ============== DataFrame 변환 ==============
def to_dataframe(annots: List[AnnotRec], comps: List[ComposedTag]):
    """AnnotRec와 ComposedTag를 DataFrame으로 변환"""

    # Equipment 추출
    equipment_annots = [a for a in annots if a.type == "equipment"]

    df_equipment = pd.DataFrame([{
        "page": a.page,
        "tag": a.text,
        "type": "equipment",
        "x1": a.bbox[0], "y1": a.bbox[1],
        "x2": a.bbox[2], "y2": a.bbox[3],
        "rgb": a.rgb,
        "hex": a.color_hex,
        "hex_swatch": a.color_hex,
        "pdf_name": a.pdf_name
    } for a in equipment_annots])

    # Instrument (composed tags) - pdf_name 포함
    df_instrument = pd.DataFrame([{
        "page": c.page,
        "tag": c.composed,
        "type": "instrument",
        "x1": c.union_bbox[0], "y1": c.union_bbox[1],
        "x2": c.union_bbox[2], "y2": c.union_bbox[3],
        "rgb": None,
        "hex": c.code_hex,
        "hex_swatch": c.code_hex,
        "pdf_name": c.pdf_name
    } for c in comps if c.number])  # 번호가 있는 것만

    # 통합
    dfs = []
    if not df_equipment.empty:
        dfs.append(df_equipment)
    if not df_instrument.empty:
        dfs.append(df_instrument)

    if dfs:
        df_result = pd.concat(dfs, ignore_index=True)
    else:
        df_result = pd.DataFrame()

    # 정렬
    if not df_result.empty:
        df_result = df_result.sort_values(["pdf_name", "page", "tag", "x1", "y1"])

    # Composed tags 상세 정보
    df_comp = pd.DataFrame([{
        "page": c.page,
        "code": c.code,
        "number": c.number,
        "composed": c.composed,
        "code_x1": c.code_bbox[0], "code_y1": c.code_bbox[1],
        "code_x2": c.code_bbox[2], "code_y2": c.code_bbox[3],
        "num_x1": c.number_bbox[0] if c.number_bbox else None,
        "num_y1": c.number_bbox[1] if c.number_bbox else None,
        "num_x2": c.number_bbox[2] if c.number_bbox else None,
        "num_y2": c.number_bbox[3] if c.number_bbox else None,
        "u_x1": c.union_bbox[0], "u_y1": c.union_bbox[1],
        "u_x2": c.union_bbox[2], "u_y2": c.union_bbox[3],
        "code_hex": c.code_hex,
        "num_hex": c.num_hex,
        "code_hex_swatch": c.code_hex,
        "num_hex_swatch": c.num_hex,
        "dy": c.dy,
        "pdf_name": c.pdf_name
    } for c in comps])

    return df_result, df_comp


# ============== 처리 워커 스레드 ==============
class ProcessWorker(QThread):
    """백그라운드 처리 스레드"""
    progress = pyqtSignal(int, int, str)
    log = pyqtSignal(str)
    finished_signal = pyqtSignal(object, object)
    error = pyqtSignal(str)

    def __init__(self, pdf_files: List[Path], output_dir: Path):
        super().__init__()
        self.pdf_files = pdf_files
        self.output_dir = output_dir
        self._is_cancelled = False

    def cancel(self):
        self._is_cancelled = True

    def run(self):
        try:
            all_annots: List[AnnotRec] = []
            all_comp: List[ComposedTag] = []

            total_files = len(self.pdf_files)

            for file_idx, pdf_path in enumerate(self.pdf_files):
                if self._is_cancelled:
                    self.log.emit("❌ 작업이 취소되었습니다.")
                    return

                self.log.emit(f"📄 처리 중: {pdf_path.name}")
                self.progress.emit(file_idx, total_files, f"PDF 처리: {pdf_path.name}")

                try:
                    doc = fitz.open(pdf_path.as_posix())
                    page_count = doc.page_count

                    for i, page in enumerate(doc):
                        if self._is_cancelled:
                            doc.close()
                            return

                        annots = collect_colored_annotations(page, exclude_gray=True, pdf_name=pdf_path.name)
                        comps = compose_vertical_pairs(annots)
                        all_annots.extend(annots)
                        all_comp.extend(comps)

                        equipment_cnt = len([a for a in annots if a.type == "equipment"])
                        instrument_cnt = len(comps)
                        self.log.emit(f"  페이지 {i+1}/{page_count}: equipment={equipment_cnt}, instrument={instrument_cnt}")

                    doc.close()
                    self.log.emit(f"✅ 완료: {pdf_path.name}")

                except Exception as e:
                    self.log.emit(f"⚠️ 오류 ({pdf_path.name}): {str(e)}")
                    continue

            self.progress.emit(total_files, total_files, "DataFrame 변환 중...")
            self.log.emit("📊 DataFrame 변환 중...")

            df_result, df_comp = to_dataframe(all_annots, all_comp)

            # 결과 저장
            self.log.emit("💾 결과 저장 중...")
            self.output_dir.mkdir(parents=True, exist_ok=True)

            csv_result = self.output_dir / "annotation_equipment_instrument_tags.csv"
            xlsx_result = self.output_dir / "annotation_equipment_instrument_tags.xlsx"
            csv_comp = self.output_dir / "annotation_instrument_details.csv"
            xlsx_comp = self.output_dir / "annotation_instrument_details.xlsx"

            # Excel 저장 전 불법 문자 제거
            df_result_clean = df_result.copy()
            df_comp_clean = df_comp.copy()

            for col in df_result_clean.select_dtypes(include=['object']).columns:
                df_result_clean[col] = df_result_clean[col].apply(sanitize_for_excel)
            for col in df_comp_clean.select_dtypes(include=['object']).columns:
                df_comp_clean[col] = df_comp_clean[col].apply(sanitize_for_excel)

            df_result_clean.to_csv(csv_result, index=False, encoding="utf-8-sig")
            df_comp_clean.to_csv(csv_comp, index=False, encoding="utf-8-sig")
            df_result_clean.to_excel(xlsx_result, index=False)
            df_comp_clean.to_excel(xlsx_comp, index=False)

            # Excel 색상 스와치
            for path, col in [(xlsx_result, "hex_swatch"),
                              (xlsx_comp, "code_hex_swatch"),
                              (xlsx_comp, "num_hex_swatch")]:
                try:
                    paint_color_swatches(path, swatch_col_name=col)
                except Exception:
                    pass

            self.log.emit(f"✅ 저장 완료:")
            self.log.emit(f"   - {csv_result}")
            self.log.emit(f"   - {xlsx_result}")
            self.log.emit(f"   - {csv_comp}")
            self.log.emit(f"   - {xlsx_comp}")

            self.finished_signal.emit(df_result, df_comp)

        except Exception as e:
            self.error.emit(str(e))


# ============== 메인 GUI ==============
class PDFAnnotationExtractorGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.pdf_files: List[Path] = []
        self.output_dir: Path = Path("out")
        self.worker: Optional[ProcessWorker] = None
        self.df_result = None
        self.df_comp = None

        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("PDF Annotation Equipment & Instrument 추출기")
        self.setMinimumSize(1000, 700)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # 상단: 파일 선택
        file_group = QGroupBox("📁 파일 설정")
        file_layout = QVBoxLayout(file_group)

        # PDF 파일 선택
        pdf_row = QHBoxLayout()
        pdf_row.addWidget(QLabel("PDF 파일:"))
        self.pdf_label = QLabel("선택된 파일 없음")
        self.pdf_label.setStyleSheet("color: gray;")
        pdf_row.addWidget(self.pdf_label, 1)
        self.btn_select_pdf = QPushButton("파일 선택")
        self.btn_select_pdf.clicked.connect(self.select_pdf_files)
        pdf_row.addWidget(self.btn_select_pdf)
        self.btn_select_folder = QPushButton("폴더 선택")
        self.btn_select_folder.clicked.connect(self.select_pdf_folder)
        pdf_row.addWidget(self.btn_select_folder)
        file_layout.addLayout(pdf_row)

        # 출력 폴더 선택
        out_row = QHBoxLayout()
        out_row.addWidget(QLabel("출력 폴더:"))
        self.out_label = QLabel(str(self.output_dir.resolve()))
        out_row.addWidget(self.out_label, 1)
        self.btn_select_out = QPushButton("변경")
        self.btn_select_out.clicked.connect(self.select_output_dir)
        out_row.addWidget(self.btn_select_out)
        file_layout.addLayout(out_row)

        main_layout.addWidget(file_group)

        # 설명
        info_group = QGroupBox("ℹ️ 정보")
        info_layout = QVBoxLayout(info_group)
        info_text = QLabel(
            "이 도구는 PDF 주석(annotation)에서 Equipment Number와 Instrument Tag를 추출합니다.\n"
            "• Equipment: AL-8903A, TK-8901 등\n"
            "• Instrument: 코드(PT, TI 등) + 번호(8903A 등) 수직 조합"
        )
        info_text.setWordWrap(True)
        info_layout.addWidget(info_text)
        main_layout.addWidget(info_group)

        # 실행 버튼
        btn_row = QHBoxLayout()
        self.btn_run = QPushButton("▶️ 실행")
        self.btn_run.setMinimumHeight(40)
        self.btn_run.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-size: 14px;
                font-weight: bold;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
        """)
        self.btn_run.clicked.connect(self.run_extraction)
        btn_row.addWidget(self.btn_run)

        self.btn_cancel = QPushButton("⏹️ 취소")
        self.btn_cancel.setMinimumHeight(40)
        self.btn_cancel.setEnabled(False)
        self.btn_cancel.clicked.connect(self.cancel_extraction)
        btn_row.addWidget(self.btn_cancel)

        main_layout.addLayout(btn_row)

        # 진행 상황
        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(True)
        main_layout.addWidget(self.progress_bar)

        self.status_label = QLabel("대기 중...")
        main_layout.addWidget(self.status_label)

        # 하단: 탭 (로그 + 결과)
        splitter = QSplitter(Qt.Orientation.Vertical)

        # 로그
        log_group = QGroupBox("📝 로그")
        log_layout = QVBoxLayout(log_group)
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setFont(QFont("Consolas", 9))
        log_layout.addWidget(self.log_text)
        splitter.addWidget(log_group)

        # 결과 탭
        result_tabs = QTabWidget()

        # Equipment & Instrument 통합 탭
        self.table_result = QTableWidget()
        self.table_result.setAlternatingRowColors(True)
        result_tabs.addTab(self.table_result, "Equipment & Instrument")

        # Instrument 상세 탭
        self.table_comp = QTableWidget()
        self.table_comp.setAlternatingRowColors(True)
        result_tabs.addTab(self.table_comp, "Instrument Details")

        splitter.addWidget(result_tabs)
        splitter.setSizes([200, 300])

        main_layout.addWidget(splitter, 1)

    def select_pdf_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self, "PDF 파일 선택", "",
            "PDF Files (*.pdf);;All Files (*)"
        )
        if files:
            self.pdf_files = [Path(f) for f in files]
            self.pdf_label.setText(f"{len(self.pdf_files)}개 파일 선택됨")
            self.pdf_label.setStyleSheet("color: black;")
            self.log_text.append(f"📄 {len(self.pdf_files)}개 파일 선택:")
            for f in self.pdf_files[:5]:
                self.log_text.append(f"   - {f.name}")
            if len(self.pdf_files) > 5:
                self.log_text.append(f"   ... 외 {len(self.pdf_files)-5}개")

    def select_pdf_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "PDF 폴더 선택")
        if folder:
            folder_path = Path(folder)
            self.pdf_files = sorted(folder_path.glob("*.pdf"))
            if self.pdf_files:
                self.pdf_label.setText(f"{len(self.pdf_files)}개 파일 (폴더: {folder_path.name})")
                self.pdf_label.setStyleSheet("color: black;")
                self.log_text.append(f"📁 폴더 선택: {folder_path}")
                self.log_text.append(f"   {len(self.pdf_files)}개 PDF 파일 발견")
            else:
                self.pdf_label.setText("선택된 폴더에 PDF 없음")
                self.pdf_label.setStyleSheet("color: red;")
                QMessageBox.warning(self, "경고", "선택된 폴더에 PDF 파일이 없습니다.")

    def select_output_dir(self):
        folder = QFileDialog.getExistingDirectory(self, "출력 폴더 선택")
        if folder:
            self.output_dir = Path(folder)
            self.out_label.setText(str(self.output_dir.resolve()))

    def run_extraction(self):
        if not self.pdf_files:
            QMessageBox.warning(self, "경고", "PDF 파일을 선택해주세요.")
            return

        self.btn_run.setEnabled(False)
        self.btn_cancel.setEnabled(True)
        self.progress_bar.setValue(0)
        self.log_text.clear()
        self.log_text.append("🚀 주석 추출 시작...")

        self.worker = ProcessWorker(
            pdf_files=self.pdf_files,
            output_dir=self.output_dir
        )

        self.worker.progress.connect(self.on_progress)
        self.worker.log.connect(self.on_log)
        self.worker.finished_signal.connect(self.on_finished)
        self.worker.error.connect(self.on_error)

        self.worker.start()

    def cancel_extraction(self):
        if self.worker:
            self.worker.cancel()
            self.btn_cancel.setEnabled(False)
            self.status_label.setText("취소 중...")

    def on_progress(self, current, total, message):
        if total > 0:
            percent = int(current / total * 100)
            self.progress_bar.setValue(percent)
        self.status_label.setText(message)

    def on_log(self, message):
        self.log_text.append(message)
        scrollbar = self.log_text.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())

    def on_finished(self, df_result, df_comp):
        self.btn_run.setEnabled(True)
        self.btn_cancel.setEnabled(False)
        self.progress_bar.setValue(100)
        self.status_label.setText("✅ 완료!")

        self.df_result = df_result
        self.df_comp = df_comp

        # 테이블 업데이트
        self.update_table(self.table_result, df_result)
        self.update_table(self.table_comp, df_comp)

        self.log_text.append("")
        self.log_text.append("=" * 50)
        self.log_text.append(f"📊 결과 요약:")
        self.log_text.append(f"   - 전체: {len(df_result)}개")

        if not df_result.empty:
            type_counts = df_result["type"].value_counts()
            self.log_text.append(f"   - 타입별 분포:")
            for t, c in type_counts.items():
                self.log_text.append(f"      {t}: {c}개")

        QMessageBox.information(self, "완료", f"추출 완료!\n\n결과 저장 위치: {self.output_dir}")

    def on_error(self, error_msg):
        self.btn_run.setEnabled(True)
        self.btn_cancel.setEnabled(False)
        self.status_label.setText("❌ 오류 발생")
        self.log_text.append(f"❌ 오류: {error_msg}")
        QMessageBox.critical(self, "오류", f"처리 중 오류 발생:\n{error_msg}")

    def update_table(self, table: QTableWidget, df: pd.DataFrame):
        if df is None or df.empty:
            table.clear()
            table.setRowCount(0)
            table.setColumnCount(0)
            return

        display_df = df.head(1000)

        table.setRowCount(len(display_df))
        table.setColumnCount(len(display_df.columns))
        table.setHorizontalHeaderLabels(display_df.columns.tolist())

        for row_idx, (_, row) in enumerate(display_df.iterrows()):
            for col_idx, value in enumerate(row):
                item = QTableWidgetItem(str(value) if pd.notna(value) else "")

                col_name = display_df.columns[col_idx]
                if "hex" in col_name.lower() and isinstance(value, str) and value.startswith("#"):
                    try:
                        color = QColor(value)
                        item.setBackground(color)
                        brightness = (color.red() * 299 + color.green() * 587 + color.blue() * 114) / 1000
                        if brightness < 128:
                            item.setForeground(QColor(255, 255, 255))
                    except:
                        pass

                table.setItem(row_idx, col_idx, item)

        table.resizeColumnsToContents()

        if len(df) > 1000:
            self.log_text.append(f"⚠️ 테이블은 처음 1000행만 표시 (전체: {len(df)}행)")


def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    window = PDFAnnotationExtractorGUI()
    window.show()

    sys.exit(app.exec())


if __name__ == "__main__":
    main()
