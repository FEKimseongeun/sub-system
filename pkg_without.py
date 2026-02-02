# -*- coding: utf-8 -*-
"""
PDF에서 주석(Polygon / Highlight / Square) 마크업 영역은 제외하고,
나머지 영역의 텍스트 span을 수집한 뒤

- 라인번호 두 줄 병합
- 코드 + 숫자 조합(composed tag) 생성
- dummy 태그 필터링
- YOLO로 instrument → special_item 재분류(선택)

까지 수행한 후

1) 기존 colored_tags / composed_tags (디버깅용)
2) 최종 결과: page, text, type, x0, y0, x1, y1 만 가진 final_tags.xlsx

를 저장하는 파이프라인.
"""
from __future__ import annotations
import csv
from pathlib import Path
from dataclasses import dataclass
from typing import List, Tuple, Dict, Optional

import time
import re
import io

import pandas as pd
import pymupdf as fitz
from shapely.geometry import Point, Polygon
from loguru import logger
from PIL import Image
import re  # 파일 최상단에 이미 있으면 생략


# ============== YOLO 모델 임포트 (선택) ==============
try:
    from ultralytics import YOLO

    YOLO_AVAILABLE = True
except ImportError:
    YOLO_AVAILABLE = False
    logger.warning("ultralytics 미설치 - YOLO 분류 비활성화")

# ============== 경로 ==============
DATA_PDF_DIR = Path("data/pdf")
OUT_DIR = Path("out")
OUT_DIR.mkdir(parents=True, exist_ok=True)

# ============== YOLO 모델 경로/파라미터 ==============
YOLO_MODEL_PATH = Path("runs/detect/symbol_detector/weights/best.pt")
CROP_MARGIN = 10
YOLO_CONFIDENCE = 0.25

CROP_IMG_DIR = Path("out/cropped_images")
CROP_IMG_DIR.mkdir(parents=True, exist_ok=True)
SAVE_CROPPED_IMAGES = True

# ============== 엑셀 색상 스와치 (기존 로직 유지) ==============
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def hex_to_argb(hex_code: str) -> str:
    if not hex_code:
        return "FFFFFFFF"
    s = hex_code.strip()
    if s.startswith("#"):
        s = s[1:]
    if len(s) != 6:
        return "FFFFFFFF"
    return "FF" + s.upper()


def make_safe_tag_for_filename(tag: str) -> str:
    """
    윈도우 파일명에 쓸 수 있도록 태그 문자열 세탁
    - 금지문자: \ / : * ? " < > |  → '_'로 치환
    - 그 외 한글/공백 등도 안전하게 쓰고 싶다면 필요시 더 줄일 수 있음
    - 너무 길어지지 않도록 50자 정도로 잘라줌
    """
    if not tag:
        return "tag"

    # 일단 문자열로 캐스팅
    s = str(tag)

    # 윈도우에서 안 되는 문자들 전부 '_'로 치환
    s = re.sub(r'[\\/:*?"<>|]', "_", s)

    # 앞뒤 공백/마침표 제거 (윈도우 파일명 끝에 . / 공백 안 됨)
    s = s.strip().rstrip(". ")

    # 너무 길면 앞부분만 사용
    if len(s) > 50:
        s = s[:50]

    # 완전 빈 문자열이 되면 fallback
    if not s:
        s = "tag"

    return s

def paint_color_swatches(xlsx_path: Path, swatch_col_name: str, header_row: int = 1):
    wb = load_workbook(xlsx_path)
    ws = wb.active
    col_idx = None
    for c in range(1, ws.max_column + 1):
        if (ws.cell(row=header_row, column=c).value or "").strip() == swatch_col_name:
            col_idx = c
            break
    if col_idx is None:
        wb.close()
        return

    for r in range(header_row + 1, ws.max_row + 1):
        hex_val = ws.cell(row=r, column=col_idx).value
        if isinstance(hex_val, str) and hex_val.strip():
            argb = hex_to_argb(hex_val)
            ws.cell(row=r, column=col_idx).fill = PatternFill(
                fill_type="solid", start_color=argb, end_color=argb
            )

    wb.save(xlsx_path)
    wb.close()


# ============== 유틸 ==============
def srgb_int_to_rgb8(srgb_int: int) -> Tuple[int, int, int]:
    r, g, b = fitz.sRGB_to_rgb(srgb_int)
    return int(r), int(g), int(b)


def rgb8_to_hex(rgb: Tuple[int, int, int]) -> str:
    r, g, b = rgb
    return f"#{r:02X}{g:02X}{b:02X}"


def bbox_center(b: Tuple[float, float, float, float]) -> Tuple[float, float]:
    x0, y0, x1, y1 = b
    return (x0 + x1) / 2.0, (y0 + y1) / 2.0


def bbox_union(a: Tuple[float, float, float, float],
               b: Tuple[float, float, float, float]) -> Tuple[float, float, float, float]:
    ax0, ay0, ax1, ay1 = a
    bx0, by0, bx1, by1 = b
    return min(ax0, bx0), min(ay0, by0), max(ax1, bx1), max(ay1, by1)


# ============== 태그 정규식 ==============
# PID NO (예: 216112C-11-PID-0021-0104)
PID_TAG_RE = re.compile(
    r'^[0-9A-Z]+-\d+-PID-\d{4}-\d{4}$',
    re.IGNORECASE,
)
# 라인넘버 (인치, 사이즈 포함)
LINE_TAG_RE = re.compile(
    r'^\d{1,3}-.*"-[A-Z]{1,3}-\d{3,6}-[A-Z0-9]{3,}-[A-Z]$',
    re.IGNORECASE,
)

# 장비 태그
EQUIP_TAG_RE = re.compile(
    r'^\d{1,3}-[A-Z]{1,4}-\d{2,5}(?:-[A-Z](?:/[A-Z])?)?$',
    re.IGNORECASE,
)

TAG_PATTERNS = {
    "line_no": re.compile(r"\b[0-9]{2,4}-[A-Z0-9]{2,}-[0-9A-Z\-]{4,}\b"),
    "valve": re.compile(r"\b(?:[A-Z]{1,3})-?\d{3,5}[A-Z]?\b"),
    "instr": re.compile(r"\b[A-Z]{1,3}-?\d{1,5}[A-Z]?\b"),
    "special": re.compile(r"\b(?:SPV|EXJ)\s?-?\s?[0-9A-Z\-]{2,}\b"),
}


def classify_tag(text: str) -> str:
    raw = (text or "").strip()
    if not raw:
        return "text"

    upper = raw.upper()

    # 1) equipment 먼저 매칭
    # 예) 11-P-621-A/B, 11-C-620, 11-PM-621-A/B
    if EQUIP_TAG_RE.fullmatch(upper):
        return "equipment"

    # 2) line (인치, 사이즈 포함 라인넘버)
    # 예) 11-6"-MW-10401-F242A-H, 11-1 1/2"-PC-13105-F400A-H
    if LINE_TAG_RE.fullmatch(upper) or TAG_PATTERNS["line_no"].search(upper):
        return "line"

    # 3) 나머지는 기존 로직 유지
    if "IF" in upper:
        return "interface"
    if TAG_PATTERNS["special"].search(upper):
        return "special"
    if TAG_PATTERNS["valve"].search(upper):
        return "valve"
    if TAG_PATTERNS["instr"].search(upper):
        return "instrument"
    return "text"


# ============== 데이터 구조 ==============
@dataclass
class SpanRec:
    page: int
    text: str
    bbox: Tuple[float, float, float, float]
    rgb: Tuple[int, int, int]
    color_hex: str
    type: str
    pdf_name: str = ""


@dataclass
class ComposedTag:
    page: int
    code: str
    number: str
    composed: str
    code_bbox: Tuple[float, float, float, float]
    number_bbox: Optional[Tuple[float, float, float, float]]
    union_bbox: Tuple[float, float, float, float]
    code_hex: str
    num_hex: Optional[str]
    dy: Optional[float]


# ============== 주석(마크업) 폴리곤 수집 ==============
def collect_markup_polygons(page: fitz.Page) -> List[Polygon]:
    polys: List[Polygon] = []
    annots = page.annots()
    if not annots:
        logger.debug("  주석 없음")
        return polys

    for annot in annots:
        polygon_coords = None

        # Polygon / Highlight 타입 (vertices 존재)
        if hasattr(annot, "vertices") and annot.vertices:
            raw_vertices = annot.vertices
            coords: List[Tuple[float, float]] = []
            for v in raw_vertices:
                # (x, y) 튜플 혹은 Point 모두 지원
                if hasattr(v, "x") and hasattr(v, "y"):
                    coords.append((float(v.x), float(v.y)))
                elif isinstance(v, (tuple, list)) and len(v) >= 2:
                    coords.append((float(v[0]), float(v[1])))
            if len(coords) >= 3:
                poly = Polygon(coords)
                polys.append(poly)
                logger.info(f"  Found {annot.type[1]} markup: {len(coords)} vertices")

        # Square 타입(rect 사용)
        elif annot.type[0] == 4:  # Square
            rect = annot.rect
            coords = [
                (rect.x0, rect.y0),
                (rect.x1, rect.y0),
                (rect.x1, rect.y1),
                (rect.x0, rect.y1),
            ]
            poly = Polygon(coords)
            polys.append(poly)
            logger.info(f"  Found Square markup: {rect}")

    logger.info(f"  Total markup regions: {len(polys)}")
    return polys


# ============== 마크업 제외 텍스트 span 수집 ==============
def collect_spans_excluding_markup(
    page: fitz.Page,
    markup_polygons: List[Polygon],
    pdf_name: str = ""
) -> List[SpanRec]:
    """
    - 페이지에서 모든 텍스트 span 수집
    - span 중심점이 markup 폴리곤 안에 있으면 제외
    - 색상은 모두 허용 (검정/회색 포함)
    """
    spans: List[SpanRec] = []
    d = page.get_text("dict")
    for blk in d.get("blocks", []):
        for line in blk.get("lines", []):
            for s in line.get("spans", []):
                text = (s.get("text") or "").strip()
                if not text:
                    continue

                x0, y0, x1, y1 = s["bbox"]
                cx, cy = bbox_center((x0, y0, x1, y1))
                pt = Point(cx, cy)

                in_markup = False
                for poly in markup_polygons:
                    if poly.contains(pt):
                        in_markup = True
                        break
                if in_markup:
                    # 마크업 영역 내부 텍스트는 제외
                    continue

                srgb_int = s.get("color")
                if srgb_int is not None:
                    rgb = srgb_int_to_rgb8(srgb_int)
                    color_hex = rgb8_to_hex(rgb)
                else:
                    rgb = (0, 0, 0)
                    color_hex = "#000000"

                spans.append(
                    SpanRec(
                        page=page.number + 1,
                        text=text,
                        bbox=(float(x0), float(y0), float(x1), float(y1)),
                        rgb=rgb,
                        color_hex=color_hex,
                        type=classify_tag(text),
                        pdf_name=pdf_name,
                    )
                )

    logger.info(f"  spans collected (excluding markup): {len(spans)}")
    return spans


# ============== 두 줄 라인넘버 병합 파라미터 ==============
LINE_PREFIX_RE = re.compile(
    r"^[0-9]{2,4}-[A-Z0-9]{2,}-[A-Z0-9\-]*-$", re.IGNORECASE
)
LINE_MERGE_DY_MAX = 25.0
LINE_MERGE_X_CENTER_TOL = 15.0


def merge_multiline_line_numbers(spans: List[SpanRec]) -> List[SpanRec]:
    """두 줄로 나뉜 라인넘버 병합 (기존 로직 유지)"""
    if not spans:
        return spans

    groups: Dict[Tuple[int, str], List[SpanRec]] = {}
    for s in spans:
        groups.setdefault((s.page, s.color_hex), []).append(s)

    merged_all: List[SpanRec] = []
    for (_, _), S in groups.items():
        S_sorted = sorted(S, key=lambda s: (s.bbox[1], s.bbox[0]))
        used = set()
        n = len(S_sorted)

        for i, s in enumerate(S_sorted):
            if i in used:
                continue

            text_top = (s.text or "").strip()
            upper_top = text_top.upper()

            if text_top.endswith("-") and LINE_PREFIX_RE.match(upper_top):
                scx, scy = bbox_center(s.bbox)
                merged_flag = False

                for j in range(i + 1, n):
                    if j in used:
                        continue
                    t = S_sorted[j]

                    dy = t.bbox[1] - s.bbox[3]
                    if dy < 0:
                        continue
                    if dy > LINE_MERGE_DY_MAX:
                        break

                    tcx, tcy = bbox_center(t.bbox)
                    if abs(tcx - scx) > LINE_MERGE_X_CENTER_TOL:
                        continue

                    text_bottom = (t.text or "").strip()
                    combined = text_top + text_bottom

                    if TAG_PATTERNS["line_no"].search(combined):
                        new_bbox = bbox_union(s.bbox, t.bbox)
                        merged_span = SpanRec(
                            page=s.page,
                            text=combined,
                            bbox=new_bbox,
                            rgb=s.rgb,
                            color_hex=s.color_hex,
                            type=classify_tag(combined),
                            pdf_name=s.pdf_name,
                        )
                        merged_all.append(merged_span)
                        used.add(i)
                        used.add(j)
                        merged_flag = True
                        break

                if not merged_flag:
                    merged_all.append(s)
            else:
                merged_all.append(s)

    return merged_all


# ============== 코드+숫자 매칭 파라미터 ==============
DX_TOL_CENTER = 11.0
DY_TOL_CENTER = 18.0

TARGET_DX = -39.1
TARGET_DY = 1.3
DX_TOL = 32.0
DY_TOL = 22.0
EXPANSIONS = [1.0, 1.5]

CODE_ONLY_RE = re.compile(r"^[A-Z]{1,4}$")
NUMBER_ONLY_RE = re.compile(r"^\d{2,5}\s*[A-Z]{0,2}$")
LETTER_ONLY_RE = re.compile(r"^[A-Z]{1,3}$")

SUFFIX_LINE_TOL = 4.0
SUFFIX_GAP_MAX  = 60.0

EXCLUDE_CODES = {"O", "L", "LL"}


def _stitch_suffix(number_span: SpanRec,
                   spans_on_page: List[SpanRec]) -> Tuple[str, Tuple[float, float, float, float]]:
    base_text = re.sub(r"\s+", "", number_span.text)
    x0, y0, x1, y1 = number_span.bbox
    ncx, ncy = bbox_center(number_span.bbox)

    suffixes: List[SpanRec] = []
    for sp in spans_on_page:
        if sp is number_span:
            continue
        if not LETTER_ONLY_RE.match(sp.text):
            continue
        scx, scy = bbox_center(sp.bbox)
        if abs(scy - ncy) > SUFFIX_LINE_TOL: continue
        if (sp.bbox[0] - x1) >= -2 and (sp.bbox[0] - x1) <= SUFFIX_GAP_MAX:
            suffixes.append(sp)

    suffixes.sort(key=lambda s: s.bbox[0])
    stitched = base_text
    stitched_bbox = (x0, y0, x1, y1)
    appended = 0
    for sp in suffixes:
        if appended >= 2:
            break
        stitched += sp.text
        sx0, sy0, sx1, sy1 = sp.bbox
        stitched_bbox = (
            min(stitched_bbox[0], sx0),
            min(stitched_bbox[1], sy0),
            max(stitched_bbox[2], sx1),
            max(stitched_bbox[3], sy1),
        )
        appended += 1

    return stitched, stitched_bbox


def _pick_by_window(numbers: List[SpanRec],
                    rx0: float, ry0: float, rx1: float, ry1: float,
                    tx: float, ty: float, cy1: float) -> Optional[Tuple[float, float, SpanRec]]:
    cand = []
    for n in numbers:
        ncx, ncy = bbox_center(n.bbox)
        if rx0 <= ncx <= rx1 and ry0 <= ncy <= ry1:
            ny0 = n.bbox[1]
            dist2 = (ncx - tx) ** 2 + (ncy - ty) ** 2
            dy = ny0 - cy1
            cand.append((dist2, dy, n))
    if not cand:
        return None
    cand.sort(key=lambda x: x[0])
    return cand[0]


def compose_vertical_pairs_simple(spans: List[SpanRec]) -> List[ComposedTag]:
    comps: List[ComposedTag] = []
    spans_by_page: Dict[int, List[SpanRec]] = {}
    for sp in spans:
        spans_by_page.setdefault(sp.page, []).append(sp)

    for page, S in spans_by_page.items():
        codes = [
            s for s in S
            if CODE_ONLY_RE.match(s.text) and s.text not in EXCLUDE_CODES
        ]
        numbers = [s for s in S if NUMBER_ONLY_RE.match(s.text)]

        for c in codes:
            cx0, cy0, cx1, cy1 = c.bbox
            ccx, _ = bbox_center(c.bbox)

            chosen = None
            tx = ccx
            ty = cy1 + 0.0
            win = (tx - DX_TOL_CENTER, ty - DY_TOL_CENTER,
                   tx + DX_TOL_CENTER, ty + DY_TOL_CENTER)
            chosen = _pick_by_window(numbers, *win, tx=tx, ty=ty, cy1=cy1)

            if not chosen:
                tx = ccx + TARGET_DX
                ty = cy1 + TARGET_DY
                for scale in EXPANSIONS:
                    win = (tx - DX_TOL * scale, ty - DY_TOL * scale,
                           tx + DX_TOL * scale, ty + DY_TOL * scale)
                    chosen = _pick_by_window(numbers, *win, tx=tx, ty=ty, cy1=cy1)
                    if chosen:
                        break

            if chosen:
                _, dy, n = chosen
                stitched_text, stitched_bbox = _stitch_suffix(n, S)
                comps.append(
                    ComposedTag(
                        page=page,
                        code=c.text,
                        number=stitched_text,
                        composed=f"{c.text}-{stitched_text}",
                        code_bbox=c.bbox,
                        number_bbox=stitched_bbox,
                        union_bbox=bbox_union(c.bbox, stitched_bbox),
                        code_hex=c.color_hex,
                        num_hex=n.color_hex,
                        dy=dy,
                    )
                )
            else:
                comps.append(
                    ComposedTag(
                        page=page,
                        code=c.text,
                        number="",
                        composed=c.text,
                        code_bbox=c.bbox,
                        number_bbox=None,
                        union_bbox=c.bbox,
                        code_hex=c.color_hex,
                        num_hex=None,
                        dy=None,
                    )
                )

    return comps


# ============== PDF 처리 (마크업 제외 + 정제) ==============
def process_pdf(pdf_path: Path):
    logger.info(f"PDF 처리 시작: {pdf_path.name}")
    doc = fitz.open(pdf_path.as_posix())
    all_spans: List[SpanRec] = []
    all_comp: List[ComposedTag] = []

    for i, page in enumerate(doc, start=1):
        t0 = time.time()
        try:
            logger.info(f"\n=== Processing Page {i}/{doc.page_count} ===")
            markup_polygons = collect_markup_polygons(page)
            spans = collect_spans_excluding_markup(page, markup_polygons, pdf_name=pdf_path.name)
            #spans = merge_multiline_line_numbers(spans)
            comps = compose_vertical_pairs_simple(spans)

            all_spans.extend(spans)
            all_comp.extend(comps)

            logger.info(
                f"  page {i}/{doc.page_count}: spans={len(spans)} "
                f"composed={len(comps)} ({time.time() - t0:.2f}s)"
            )
        except Exception as e:
            logger.exception(f"  page {i} 오류: {e} -> 건너뜀")
            continue

    doc.close()
    logger.info(
        f"PDF 완료: {pdf_path.name} "
        f"(spans={len(all_spans)}, composed={len(all_comp)})"
    )
    return all_spans, all_comp


def to_dataframe(spans: List[SpanRec], comps: List[ComposedTag]):

    df_spans = pd.DataFrame([{
        "page": s.page,
        "tag": s.text,
        "type": s.type,
        "x1": s.bbox[0], "y1": s.bbox[1],
        "x2": s.bbox[2], "y2": s.bbox[3],
        "rgb": s.rgb,
        "hex": s.color_hex,
        "hex_swatch": s.color_hex,
        "pdf_name": s.pdf_name
    } for s in spans])

    df_comp = pd.DataFrame([{
        "page": c.page,
        "code": c.code,
        "number": c.number,
        "composed": c.composed,

        "code_x1": c.code_bbox[0], "code_y1": c.code_bbox[1],
        "code_x2": c.code_bbox[2], "code_y2": c.code_bbox[3],

        "num_x1": c.number_bbox[0] if c.number_bbox else None,
        "num_y1": c.number_bbox[1] if c.number_bbox else None,
        "num_x2": c.number_bbox[2] if c.number_bbox else None,
        "num_y2": c.number_bbox[3] if c.number_bbox else None,

        "u_x1": c.union_bbox[0], "u_y1": c.union_bbox[1],
        "u_x2": c.union_bbox[2], "u_y2": c.union_bbox[3],

        "code_hex": c.code_hex,
        "num_hex": c.num_hex,

        "code_hex_swatch": c.code_hex,
        "num_hex_swatch": c.num_hex,

        "dy": c.dy
    } for c in comps])

    if not df_spans.empty and not df_comp.empty:
        df_comp_nonempty = df_comp[df_comp["number"].astype(str) != ""]
        used_indices = set()

        def _center_dist2(b1, b2):
            c1x, c1y = bbox_center(b1)
            c2x, c2y = bbox_center(b2)
            return (c1x - c2x) ** 2 + (c1y - c2y) ** 2

        for _, row in df_comp_nonempty.iterrows():
            page = row["page"]
            code = row["code"]
            composed = row["composed"]
            code_hex = row["code_hex"]

            # 후보: 같은 page + 같은 code 텍스트 + 같은 색(hex)
            candidates = df_spans[
                (df_spans["page"] == page) &
                (df_spans["tag"] == code) &
                (df_spans["hex"] == code_hex)
            ]

            if candidates.empty:
                continue

            # 이미 다른 composed가 사용한 span 제외
            candidates = candidates[~candidates.index.isin(used_indices)]
            if candidates.empty:
                continue

            # code_bbox와 가장 가까운 span 하나 선택
            code_bbox = (row["code_x1"], row["code_y1"],
                         row["code_x2"], row["code_y2"])

            dist2 = candidates.apply(
                lambda r: _center_dist2(
                    (r["x1"], r["y1"], r["x2"], r["y2"]),
                    code_bbox
                ),
                axis=1
            )

            best_idx = dist2.idxmin()
            used_indices.add(best_idx)

            # 해당 span 1개만 composed로 치환
            df_spans.loc[best_idx, "tag"] = composed
            df_spans.loc[best_idx, "type"] = "instrument"
            df_spans.loc[best_idx, ["x1", "y1", "x2", "y2"]] = [
                row["u_x1"], row["u_y1"], row["u_x2"], row["u_y2"]
            ]

    if not df_spans.empty:
        df_spans["tag"] = df_spans["tag"].astype(str)

        # (1) 제외 코드(O, L, LL)
        mask_exclude_codes = df_spans["tag"].isin(EXCLUDE_CODES)

        # (2) 숫자만 있는 값
        mask_digits_only = df_spans["tag"].str.fullmatch(r"\d+")

        # (3) 알파+숫자만 붙은 단어 (7302A, A4500, N05 등)
        upper_tags = df_spans["tag"].str.upper()
        mask_alnum_word = upper_tags.str.fullmatch(r"(?=.*[A-Z])(?=.*\d)[A-Z0-9]+")

        df_spans = df_spans[~(mask_exclude_codes | mask_digits_only | mask_alnum_word)].copy()

        # 정렬 후 (tag, page, hex) 단위 중복 제거
        df_spans = (
            df_spans
            .sort_values(["tag", "page", "hex", "x1", "y1"])
            .drop_duplicates(["tag", "page", "hex"], keep="first")
        )

    return df_spans, df_comp



# ============== YOLO 모델 로드 ==============
def load_yolo_model(model_path: Path) -> Optional["YOLO"]:
    if not YOLO_AVAILABLE:
        logger.warning("YOLO 미사용 - ultralytics 미설치")
        return None

    if not model_path.exists():
        logger.warning(f"YOLO 모델 없음: {model_path} - 자동 분류 건너뜀")
        return None

    try:
        model = YOLO(model_path.as_posix())
        logger.info(f"✅ YOLO 모델 로드 완료: {model_path}")
        return model
    except Exception as e:
        logger.error(f"YOLO 모델 로드 실패: {e}")
        return None


# ============== PDF 좌표로 이미지 Crop ==============
def crop_bbox_from_pdf(
    pdf_path: Path,
    page_num: int,
    bbox: Tuple[float, float, float, float],
    margin: int = CROP_MARGIN,
) -> Optional[Image.Image]:
    """
    PDF에서 특정 bbox 영역을 잘라 800x800 패딩된 이미지를 반환
    """
    try:
        doc = fitz.open(pdf_path.as_posix())
        page = doc[page_num - 1]

        x0, y0, x1, y1 = bbox

        x0 = max(0, x0 - margin)
        y0 = max(0, y0 - margin)
        x1 = min(page.rect.width, x1 + margin)
        y1 = min(page.rect.height, y1 + margin)

        clip_rect = fitz.Rect(x0, y0, x1, y1)
        pix = page.get_pixmap(clip=clip_rect, matrix=fitz.Matrix(2.0, 2.0))

        import cv2
        import numpy as np

        img_data = pix.tobytes("png")
        img_pil = Image.open(io.BytesIO(img_data))
        img_array = np.array(img_pil)

        # RGB → BGR
        if len(img_array.shape) == 3 and img_array.shape[2] == 3:
            img_cv = cv2.cvtColor(img_array, cv2.COLOR_RGB2BGR)
        else:
            img_cv = img_array

        tile_size = 800
        h, w = img_cv.shape[:2]

        # 작은 경우 패딩
        if h < tile_size or w < tile_size:
            img_cv = cv2.copyMakeBorder(
                img_cv,
                0,
                max(0, tile_size - h),
                0,
                max(0, tile_size - w),
                cv2.BORDER_CONSTANT,
                value=[255, 255, 255],
            )
        # 큰 경우 중앙 크롭 (+ 필요시 재패딩)
        elif h > tile_size or w > tile_size:
            center_y, center_x = h // 2, w // 2
            y0c = max(0, center_y - tile_size // 2)
            x0c = max(0, center_x - tile_size // 2)
            y1c = min(h, y0c + tile_size)
            x1c = min(w, x0c + tile_size)
            img_cv = img_cv[y0c:y1c, x0c:x1c]

            h_new, w_new = img_cv.shape[:2]
            if h_new < tile_size or w_new < tile_size:
                img_cv = cv2.copyMakeBorder(
                    img_cv,
                    0,
                    max(0, tile_size - h_new),
                    0,
                    max(0, tile_size - w_new),
                    cv2.BORDER_CONSTANT,
                    value=[255, 255, 255],
                )

        img_rgb = cv2.cvtColor(img_cv, cv2.COLOR_BGR2RGB)
        img_pil_final = Image.fromarray(img_rgb)

        logger.debug(f"    Crop 이미지 크기: {img_pil_final.size}")

        doc.close()
        return img_pil_final

    except Exception as e:
        logger.error(f"이미지 crop 실패 (page {page_num}, bbox {bbox}): {e}")
        return None


# ============== YOLO로 special_item 판정 ==============
def classify_with_yolo(
    model: "YOLO",
    img: Image.Image,
    confidence: float = YOLO_CONFIDENCE,
) -> bool:
    try:
        results = model(img, conf=confidence, imgsz=800, verbose=False)

        if len(results) == 0 or len(results[0].boxes) == 0:
            return False

        for box in results[0].boxes:
            class_id = int(box.cls[0])
            if class_id == 0:  # special_item
                return True

        return False

    except Exception as e:
        logger.error(f"YOLO 추론 실패: {e}")
        return False


# ============== instrument → special_item 재분류 ==============
def reclassify_instruments_with_yolo(
    df_spans: pd.DataFrame,
    pdf_dir: Path,
    model: Optional["YOLO"],
) -> pd.DataFrame:
    if model is None:
        logger.info("YOLO 모델 없음 - 재분류 스킵")
        return df_spans

    if df_spans.empty:
        return df_spans

    instruments = df_spans[df_spans["type"] == "instrument"].copy()
    if instruments.empty:
        logger.info("instrument 타입 없음 - 재분류 스킵")
        return df_spans

    logger.info(f"🔍 YOLO 재분류 시작: {len(instruments)}개 instrument 검사")

    reclassified_indices = []
    processed = 0

    preprocessed_dir = CROP_IMG_DIR / "preprocessed_800x800"
    preprocessed_dir.mkdir(exist_ok=True)

    import cv2
    import numpy as np

    for idx, row in instruments.iterrows():
        pdf_name = row["pdf_name"]
        pdf_path = pdf_dir / pdf_name

        if not pdf_path.exists():
            logger.warning(f"PDF 없음: {pdf_path}")
            continue

        page_num = int(row["page"])
        bbox = (row["x1"], row["y1"], row["x2"], row["y2"])
        tag = row["tag"]

        img = crop_bbox_from_pdf(pdf_path, page_num, bbox, margin=CROP_MARGIN)
        if img is None:
            continue

        # ★ 파일명 안전하게 세탁
        safe_tag = make_safe_tag_for_filename(tag)
        img_filename = f"{Path(pdf_name).stem}_p{page_num}_{safe_tag}_{idx}.jpg"

        if SAVE_CROPPED_IMAGES:
            import cv2
            import numpy as np

            # 원본 crop 저장
            img.save(CROP_IMG_DIR / img_filename, "JPEG", quality=95)

            # 전처리 후 800x800 저장
            img_array = np.array(img)
            if len(img_array.shape) == 3:
                img_cv = cv2.cvtColor(img_array, cv2.COLOR_RGB2BGR)
            else:
                img_cv = img_array

            target_size = 800
            h, w = img_cv.shape[:2]
            scale = min(target_size / w, target_size / h)
            new_w, new_h = int(w * scale), int(h * scale)
            resized = cv2.resize(img_cv, (new_w, new_h))

            padded = np.full((target_size, target_size, 3), 255, dtype=np.uint8)
            y_offset = (target_size - new_h) // 2
            x_offset = (target_size - new_w) // 2
            padded[y_offset:y_offset + new_h, x_offset:x_offset + new_w] = resized

            cv2.imwrite(str(preprocessed_dir / img_filename), padded)

        # YOLO 판정
        is_special = classify_with_yolo(model, img, confidence=YOLO_CONFIDENCE)
        if is_special:
            reclassified_indices.append(idx)
            logger.debug(f"  ✅ special_item 발견: {tag} (page {page_num})")

        processed += 1
        if processed % 50 == 0:
            logger.info(
                f"  진행: {processed}/{len(instruments)} "
                f"({len(reclassified_indices)} special_item)"
            )

    if reclassified_indices:
        df_spans.loc[reclassified_indices, "type"] = "special_item"
        logger.info(
            f"✨ 재분류 완료: {len(reclassified_indices)}개 "
            f"instrument → special_item"
        )
    else:
        logger.info("재분류 결과: special_item 없음")

    if SAVE_CROPPED_IMAGES:
        saved_count = len(list(CROP_IMG_DIR.glob("*.jpg")))
        logger.info(
            f"💾 Cropped 이미지 저장 완료: {saved_count}개 → {CROP_IMG_DIR}"
        )

    return df_spans

def attach_pid_no_by_page(df_spans: pd.DataFrame) -> pd.DataFrame:
    """
    각 page마다 PID_TAG_RE에 매칭되는 태그를 찾아서
    'PID NO' 컬럼으로 붙여준다.
    - 페이지당 1개라고 가정
    - 0개면 None
    - 2개 이상이면 첫 번째만 쓰고 warning
    """
    if df_spans.empty:
        df_spans["PID NO"] = None
        return df_spans

    pid_by_page = {}

    for page, grp in df_spans.groupby("page"):
        tags = grp["tag"].astype(str)

        candidates = tags[tags.str.match(PID_TAG_RE, na=False)]

        if len(candidates) == 1:
            pid_val = candidates.iloc[0]
        elif len(candidates) > 1:
            pid_val = candidates.iloc[0]
            logger.warning(
                f"page {page}: PID 후보가 {len(candidates)}개 → 첫 번째만 사용: {pid_val}"
            )
        else:
            pid_val = None
            logger.warning(f"page {page}: PID 패턴 매칭 없음")

        pid_by_page[page] = pid_val

    df_spans["PID NO"] = df_spans["page"].map(pid_by_page)
    return df_spans

# ============== 메인 ==============
def main():
    logger.add(OUT_DIR / "pdf_color_extract.log", rotation="500 KB")

    pdf_list = sorted(DATA_PDF_DIR.glob("*.pdf"))
    if not pdf_list:
        logger.error(f"PDF가 없습니다: {DATA_PDF_DIR.resolve()}")
        return

    # 1) PDF들 처리 → SpanRec / ComposedTag 수집
    Gs: List[SpanRec] = []
    Gc: List[ComposedTag] = []
    for pdf in pdf_list:
        try:
            s, c = process_pdf(pdf)
            Gs.extend(s)
            Gc.extend(c)
        except Exception as e:
            logger.exception(f"{pdf.name} 처리 오류: {e} -> 건너뜀")
            continue

    # 2) DataFrame 변환 + composed 반영 + 정제
    df_spans, df_comp = to_dataframe(Gs, Gc)

    # 3) YOLO로 instrument → special_item 재분류

    model = load_yolo_model(YOLO_MODEL_PATH)
    df_spans = reclassify_instruments_with_yolo(df_spans, DATA_PDF_DIR, model)
    df_spans = attach_pid_no_by_page(df_spans)
    df_spans = df_spans[df_spans["type"] != "text"].copy()

    # 4) 기존 결과 저장 (colored_tags / composed_tags)
    xlsx_spans = OUT_DIR / "colored_tags_without_pkg.xlsx"
    xlsx_comp = OUT_DIR / "colored_tags_without_pkg_com.xlsx"

    df_spans.to_excel(xlsx_spans, index=False)
    df_comp.to_excel(xlsx_comp, index=False)

    # 색상 스와치 (기존 로직 유지)
    for path, col in [
        (xlsx_spans, "hex_swatch"),
        (xlsx_comp, "code_hex_swatch"),
        (xlsx_comp, "num_hex_swatch"),
    ]:
        try:
            paint_color_swatches(path, swatch_col_name=col)
        except Exception as e:
            logger.warning(f"{path.name} 스와치 경고: {e}")

    logger.info(f"SAVED → {xlsx_spans},  {xlsx_comp}")

    # 5) ✅ 최종 결과: page, text, type, x0, y0, x1, y1 형식으로 별도 저장
    if not df_spans.empty:
        df_final = (
            df_spans.rename(
                columns={
                    "tag": "text",
                    "x1": "x0",
                    "y1": "y0",
                    "x2": "x1",
                    "y2": "y1",
                }
            )[["page", "PID NO", "text", "type", "x0", "y0", "x1", "y1"]]
            .reset_index(drop=True)
        )
        df_final = df_final[df_final["type"] != "text"].reset_index(drop=True)

        final_xlsx = OUT_DIR / "final_tags.xlsx"
        df_final.to_excel(final_xlsx, index=False)

        logger.info(
            f"✅ 최종 결과 저장 → {final_xlsx} "
            f"(cols: page, PID NO, text, type, x0, y0, x1, y1)"
        )


if __name__ == "__main__":
    main()
